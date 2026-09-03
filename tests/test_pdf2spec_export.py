"""Unit tests for src.pdf2spec.export_xlsx — XLSX generation."""
import tempfile
from pathlib import Path

import pytest
from src.pdf2spec.export_xlsx import export_xlsx, HDR_VK, HDR_OV


class TestExportXlsx:
    def _make_row(self, **kwargs):
        defaults = {
            'role': 'item', 'name': 'Труба', 'type': 'DN15', 'code': '',
            'supplier': 'Завод', 'unit': 'м', 'qty': '10', 'mass': '',
            'note': '', 'poz': '1',
        }
        defaults.update(kwargs)
        return defaults

    def test_creates_xlsx(self):
        rows = [self._make_row()]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'test.xlsx'
            result = export_xlsx(rows, path)
            assert result.exists()
            assert result.stat().st_size > 0

    def test_ov_headers(self):
        rows = [self._make_row()]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'test.xlsx'
            export_xlsx(rows, path, template='OV')
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            assert ws.cell(1, 1).value == 'Позиция'
            assert ws.cell(1, 2).value == 'Наименование и техническая характеристика'

    def test_vk_headers(self):
        rows = [self._make_row()]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'test.xlsx'
            export_xlsx(rows, path, template='VK')
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            assert ws.cell(1, 1).value == 'Поз.'

    def test_header_row(self):
        rows = [self._make_row(role='header', name='АРМАТУРА', type='', qty='')]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'test.xlsx'
            export_xlsx(rows, path)
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            assert ws.cell(2, 2).value == 'АРМАТУРА'
            assert ws.cell(2, 7).value in ('', None)  # no qty for headers

    def test_component_row(self):
        rows = [self._make_row(role='component', name='а) Сифон',
                               type='ГОСТ 23289-94', qty='', supplier='Завод')]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'test.xlsx'
            export_xlsx(rows, path)
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            assert ws.cell(2, 3).value == 'ГОСТ 23289-94'
            assert ws.cell(2, 5).value == 'Завод'

    def test_cyrillic_intact(self):
        rows = [self._make_row(name='Труба стальная Ø150х4,5')]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'test.xlsx'
            export_xlsx(rows, path)
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            assert 'Ø150' in ws.cell(2, 2).value

    def test_multiple_rows(self):
        rows = [
            self._make_row(poz='1', name='Труба', qty='10'),
            self._make_row(poz='2', name='Кран', qty='5'),
            self._make_row(poz='3', name='Фитинг', qty='20'),
        ]
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / 'test.xlsx'
            export_xlsx(rows, path)
            import openpyxl
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            assert ws.max_row == 4  # header + 3 data rows
