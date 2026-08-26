"""Тесты ExcelWriter (Qt-free, openpyxl) — загрузка спецификаций, детект колонок, запись результатов."""

from pathlib import Path

import openpyxl
import pytest

from src.excel_writer import ExcelWriter, SpecItem, _clean_brand


def make_spec_xlsx(tmp_path, headers, rows):
    path = tmp_path / "spec.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(path)
    return str(path)


@pytest.fixture
def writer(tmp_path):
    path = make_spec_xlsx(
        tmp_path,
        ["№", "Наименование", "Марка", "Артикул", "Кол-во", "Ед."],
        [
            [1, "Кабель ВВГ", "Спецкабель", "ВВГ-3x2.5", 100, "м"],
            [2, "Труба ПНД", "ОАО", "", 50, "шт"],
        ],
    )
    w = ExcelWriter({"paths": {"data_output": str(tmp_path / "out")}})
    w.load_spec(path)
    return w


class TestSpecItem:
    def test_defaults(self):
        item = SpecItem("text")
        assert item.text == "text"
        assert item.article == ""
        assert item.brand == ""
        assert item.uom == "шт"
        assert item.headers == []

    def test_full(self):
        item = SpecItem("t", article="a", brand="b", name_raw="n", uom="м", headers=["h"])
        assert item.article == "a"
        assert item.uom == "м"


class TestLoadSpec:
    def test_counts_data_rows(self, writer):
        assert writer.total_rows == 2

    def test_headers_loaded(self, writer):
        assert writer.headers[1] == "Наименование"
        assert writer.headers[4] == "Кол-во"

    def test_properties(self, writer):
        assert writer.ws is not None
        assert writer.header_map is not None
        assert writer.spec_path.endswith(".xlsx")

    def test_empty_workbook_properties(self):
        w = ExcelWriter({})
        assert w.headers == []
        assert w.ws is None
        assert w.spec_path is None
        assert w.total_rows == 0


class TestDetectColumns:
    def test_typical(self):
        w = ExcelWriter({})
        mapping = w.detect_columns(["Наименование", "Артикул", "Марка", "Кол-во", "Ед."])
        assert mapping["name"] == [0]
        assert mapping["article"] == [1]
        assert mapping["brand"] == [2]
        assert mapping["qty"] == 3
        assert mapping["uom"] == 4

    def test_brand_not_name(self):
        w = ExcelWriter({})
        mapping = w.detect_columns(["Наименование", "Производитель"])
        assert mapping["brand"] == [1]
        assert mapping["name"] == [0]

    def test_no_name_fallback(self):
        w = ExcelWriter({})
        mapping = w.detect_columns(["Артикул", "Цена"])
        assert mapping["name"]

    def test_none_header_skipped(self):
        w = ExcelWriter({})
        mapping = w.detect_columns([None, "Кол-во"])
        assert mapping["name"] == []


class TestBuildItemName:
    def test_combines_brand_and_name(self, writer):
        mapping = writer.detect_columns(writer.headers)
        name, uom, article = writer.build_item_name(2, mapping)
        # производитель НЕ конкатенируется в имя (держится отдельно);
        # артикул (ВВГ-3x2.5) — участвует наравне
        assert name == "Кабель ВВГ ВВГ-3x2.5"
        assert uom == "м"
        assert article == "ВВГ-3x2.5"
        brand = writer._concat_cells(2, mapping.get("brand", []))
        assert brand == "Спецкабель"

    def test_no_ws(self):
        w = ExcelWriter({})
        assert w.build_item_name(1, {"article": [], "brand": [], "name": [], "uom": None}) == ("", "шт", None)

    def test_spec_column_appended_to_name(self, tmp_path):
        """«Тип, марка» — различитель типоразмеров — обязана попасть в имя."""
        path = make_spec_xlsx(
            tmp_path,
            ["Наименование", "Тип, марка, обозначение документа", "Кол-во", "Ед."],
            [["Стальной панельный радиатор LEMAX", "LEMAX Premium C10 500x400", 3, "шт."]],
        )
        w = ExcelWriter({})
        w.load_spec(path)
        mapping = w.detect_columns(w.headers)
        name, _, _ = w.build_item_name(2, mapping)
        assert name == "Стальной панельный радиатор LEMAX LEMAX Premium C10 500x400"

    def test_standard_reference_not_appended(self, tmp_path):
        """ГОСТ/ТУ/СТО — не модель товара, в поисковое имя не добавляется."""
        path = make_spec_xlsx(
            tmp_path,
            ["Наименование", "Тип, марка, обозначение документа", "Кол-во", "Ед."],
            [["Трубка ENERGOFLEX Super SK 60/40-2", "ГОСТ Р 56729-2015", 40, "м"]],
        )
        w = ExcelWriter({})
        w.load_spec(path)
        mapping = w.detect_columns(w.headers)
        name, _, _ = w.build_item_name(2, mapping)
        assert name == "Трубка ENERGOFLEX Super SK 60/40-2"

    def test_duplicate_value_not_appended(self, tmp_path):
        path = make_spec_xlsx(
            tmp_path,
            ["Наименование", "Марка", "Кол-во", "Ед."],
            [["Клапан RTR-G угловой", "RTR-G угловой", 27, "шт."]],
        )
        w = ExcelWriter({})
        w.load_spec(path)
        # «Марка» здесь классифицируется как brand; проверяем защиту от дублей напрямую
        name, _, _ = w.build_item_name(2, {"article": [], "brand": [], "name": [0], "spec": [1], "uom": 3})
        assert name.count("RTR-G") == 1

    def test_article_participates_in_name(self, tmp_path):
        """«Код оборудования» — идентификатор товара — входит в spec_text наравне с типом."""
        path = make_spec_xlsx(
            tmp_path,
            ["Наименование", "Тип, марка, обозначение документа", "Код оборудования", "Кол-во", "Ед."],
            [["Кран шаровой полнопроходной латунный никелированный с накидной гайкой",
              "DN15", "065B8203R", 196, "шт."]],
        )
        w = ExcelWriter({})
        w.load_spec(path)
        mapping = w.detect_columns(w.headers)
        name, _, article = w.build_item_name(2, mapping)
        assert name == "Кран шаровой полнопроходной латунный никелированный с накидной гайкой DN15 065B8203R"
        assert article == "065B8203R"

    def test_energoflex_triple_column_absorbed(self, tmp_path):
        """ENERGOFLEX в имени + ГОСТ в типе + ENERGOFLEX в коде — одно упоминание (скриншот)."""
        path = make_spec_xlsx(
            tmp_path,
            ["Наименование", "Тип, марка, обозначение документа", "Код оборудования", "Кол-во", "Ед."],
            [["Трубка ENERGOFLEX Super SK 60/40-2", "ГОСТ Р 56729-2015", "ENERGOFLEX", 40.7, "м"]],
        )
        w = ExcelWriter({})
        w.load_spec(path)
        mapping = w.detect_columns(w.headers)
        name, _, _ = w.build_item_name(2, mapping)
        assert name.lower().count("energoflex") == 1

    def test_tr84_vs_tr_84_normalized_absorbed(self, tmp_path):
        """«TR 84» в типе поглощается «TR84» из имени несмотря на пробел."""
        path = make_spec_xlsx(
            tmp_path,
            ["Наименование", "Тип, марка, обозначение документа", "Кол-во", "Ед."],
            [["Термостатический элемент TR84 с датчиком", "TR 84", 196, "шт."]],
        )
        w = ExcelWriter({})
        w.load_spec(path)
        mapping = w.detect_columns(w.headers)
        name, _, _ = w.build_item_name(2, mapping)
        assert name == "Термостатический элемент TR84 с датчиком"
        assert name.lower().count("tr84") == 1


class TestIdentAbsorption:
    """Нормализованное поглощение повторяющихся данных из разных колонок."""

    def test_short_numeric_not_absorbed_in_number(self):
        # «100» не поглощается внутри «500x1000»
        from src.excel_writer import _value_absorbed
        assert _value_absorbed("Труба 500x1000", "100") is False

    def test_long_code_substring_absorbed(self):
        from src.excel_writer import _value_absorbed
        assert _value_absorbed("Кабель ВВГ 3x1.5", "ВВГ 3x1.5") is True

    def test_separator_variant_absorbed(self):
        from src.excel_writer import _value_absorbed
        assert _value_absorbed("Трубка ENERGOFLEX Super SK 60/40-2", "ENERGOFLEX") is True

    def test_partial_overlap_not_absorbed(self):
        from src.excel_writer import _value_absorbed
        assert _value_absorbed(
            "Стальной панельный радиатор LEMAX Premium Compact Hygiene",
            "LEMAX Premium C10 500x400",
        ) is False


class TestGetSpecs:
    def test_builds_specs_skipping_empty(self, tmp_path):
        path = make_spec_xlsx(tmp_path, ["Наименование"], [["Кабель"], [], ["Труба"]])
        w = ExcelWriter({})
        w.load_spec(path)
        specs = w.get_specs()
        assert len(specs) == 2
        assert specs[0].text == "Кабель"
        assert specs[0].uom == "шт"

    def test_empty_ws(self):
        w = ExcelWriter({})
        assert w.get_specs() == []

    def test_skips_section_headers_without_qty(self, tmp_path):
        """Строки-заголовки разделов (без количества) не становятся товарами."""
        path = make_spec_xlsx(
            tmp_path,
            ["Позиция", "Наименование", "Кол-во"],
            [
                ["", "Отопление", ""],
                ["1.", "Воздухоотводчик", "48"],
                ["2.", "Кран шаровой", "18"],
            ],
        )
        w = ExcelWriter({})
        w.load_spec(path)
        specs = w.get_specs()
        assert [s.text for s in specs] == ["Воздухоотводчик", "Кран шаровой"]

    def test_row_points_to_real_sheet_row(self, tmp_path):
        """SpecItem.row — фактический ряд листа, даже после пропущенных строк-заголовков.

        Регрессия: запись результатов по «индекс в отфильтрованном списке + 2»
        смещала все цены на чужие строки (файл vtk_spec_v2.xlsx).
        """
        path = make_spec_xlsx(
            tmp_path,
            ["Позиция", "Наименование", "Кол-во"],
            [
                ["", "Отопление", ""],          # sheet row 2 — заголовок, пропускается
                ["1.", "Воздухоотводчик", "48"],  # sheet row 3
                ["", "Вентиляция", ""],          # sheet row 4 — заголовок
                ["2.", "Кран шаровой", "18"],     # sheet row 5
            ],
        )
        w = ExcelWriter({})
        w.load_spec(path)
        specs = w.get_specs()
        assert [(s.text, s.row) for s in specs] == [
            ("Воздухоотводчик", 3),
            ("Кран шаровой", 5),
        ]

    def test_no_qty_column_keeps_all(self, tmp_path):
        path = make_spec_xlsx(tmp_path, ["Наименование"], [["Кабель"], ["Труба"]])
        w = ExcelWriter({})
        w.load_spec(path)
        specs = w.get_specs()
        assert len(specs) == 2


class TestSpecForRow:
    def test_matches_get_specs(self, tmp_path):
        """spec_for_row для строки равен SpecItem из get_specs — отметки совпадут 1:1."""
        path = make_spec_xlsx(
            tmp_path,
            ["Наименование", "Марка", "Артикул", "Кол-во", "Ед."],
            [
                ["Кабель ВВГ", "Спецкабель", "ВВГ-3x2.5", 100, "м"],
                ["Труба ПНД", "ОАО", "", 50, "шт"],
            ],
        )
        w = ExcelWriter({})
        w.load_spec(path)
        specs = w.get_specs()
        assert len(specs) == 2
        row2 = w.spec_for_row(2)
        row3 = w.spec_for_row(3)
        assert row2 is not None and row3 is not None
        assert row2.text == specs[0].text == "Кабель ВВГ ВВГ-3x2.5"
        assert row2.brand == specs[0].brand == "Спецкабель"
        assert row3.text == specs[1].text == "Труба ПНД"
        assert row3.article == specs[1].article == ""

    def test_skips_section_header(self, tmp_path):
        path = make_spec_xlsx(
            tmp_path,
            ["Наименование", "Кол-во"],
            [["Отопление", ""], ["Воздухоотводчик", "48"]],
        )
        w = ExcelWriter({})
        w.load_spec(path)
        assert w.spec_for_row(2) is None
        assert w.spec_for_row(3) is not None

    def test_skips_empty_name(self, tmp_path):
        path = make_spec_xlsx(tmp_path, ["Наименование"], [[""], ["Кран"]])
        w = ExcelWriter({})
        w.load_spec(path)
        assert w.spec_for_row(2) is None
        assert w.spec_for_row(3) is not None

    def test_no_ws(self):
        w = ExcelWriter({})
        assert w.spec_for_row(2) is None


class TestCleanBrand:
    def test_strips_quotes(self):
        assert _clean_brand('"Ридан"') == "Ридан"
        assert _clean_brand('«SPL»') == "SPL"

    def test_drops_country_only(self):
        for v in ["Россия", '"Россия"', "РФ", "Российская Федерация"]:
            assert _clean_brand(v) == "", f"brand not dropped: {v!r}"

    def test_keeps_real_brand(self):
        assert _clean_brand("Арктос (Россия)") == "Арктос (Россия)"
        assert _clean_brand("Aerostar") == "Aerostar"

    def test_empty(self):
        assert _clean_brand("") == ""
        assert _clean_brand("  ") == ""


class TestWriteAndSave:
    def test_output_headers_added(self, writer):
        hm = writer.header_map
        assert writer.ws.cell(1, hm["price"]).value == "Цена, RUB"
        assert writer.ws.cell(1, hm["url"]).value == "URL"
        assert writer.ws.cell(1, hm["category"]).value == "Категория"

    def test_existing_output_headers_reused(self, tmp_path):
        path = make_spec_xlsx(tmp_path, ["Наименование", "Цена, RUB"], [["X", 100]])
        w = ExcelWriter({})
        w.load_spec(path)
        hm = w.header_map
        assert hm["price"] == 2

    def test_write_result(self, writer):
        writer.write_result(2, {
            "final_price_rub": 1500.5, "card_url": "http://x.ru",
            "primary_cat": "cables", "primary_subcat": "power",
        })
        hm = writer.header_map
        assert writer.ws.cell(2, hm["price"]).value == 1500.5
        assert writer.ws.cell(2, hm["url"]).value == "http://x.ru"
        assert "power" in writer.ws.cell(2, hm["category"]).value

    def test_write_result_no_state(self, writer):
        writer.write_result(2, {})
        hm = writer.header_map
        assert writer.ws.cell(2, hm["price"]).value is None

    def test_write_result_no_ws(self):
        w = ExcelWriter({})
        w.write_result(1, {"final_price_rub": 1})  # no crash

    def test_save_output_copy(self, writer, tmp_path):
        out_path = writer.save_output_copy(str(tmp_path / "out"))
        assert out_path.endswith(".xlsx")
        assert Path(out_path).exists()

    def test_save_output_copy_no_path(self, tmp_path):
        w = ExcelWriter({})
        assert w.save_output_copy(str(tmp_path)) == ""

    def test_flush(self, writer, tmp_path):
        writer.flush()
        out = list((tmp_path / "out").glob("*.xlsx"))
        assert len(out) >= 1
