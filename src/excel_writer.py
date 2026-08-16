"""
ExcelWriter — единый класс для всей работы с Excel.

Объединяет:
- Загрузку XLSX-спецификаций
- Определение колонок (name/article/brand/uom/qty)
- Поиск/создание выходных колонок (price/URL/category)
- Сборку имени товара из строки Excel
- Запись результатов
- Сохранение в timestamped output-файл (оригинал НЕ трогает)
"""

import logging
import threading
from pathlib import Path
from datetime import datetime
from typing import Optional

import openpyxl


logger = logging.getLogger(__name__)


class SpecItem:
    """Структурированное описание товара из Excel."""
    def __init__(self, text: str, article: str = "", brand: str = "", name_raw: str = "",
                 uom: str = "шт", headers: list | None = None, spec: str = ""):
        self.text = text
        self.article = article
        self.brand = brand
        self.name_raw = name_raw
        self.uom = uom
        self.headers = headers or []
        self.spec = spec


from src._labels import _CAT_RU_LABELS, _SUBCAT_RU_LABELS
from src.column_classifier import classify_columns


logger = logging.getLogger(__name__)

# Значения колонки «Завод-изготовитель», которые НЕ являются брендом
_COUNTRY_ONLY = {
    "россия", "рф", "российская федерация", "снг", "россия (рф)",
    "казахстан", "беларусь", "украина", "китай", "кнр", "германия",
    "италия", "польша", "турция", "япония", "сша", "швеция", "финляндия",
}


def _clean_brand(value: str) -> str:
    """Очищает значение производителя: кавычки и «страновые» значения."""
    v = value.strip().strip('"«»')
    if not v:
        return ""
    if v.lower() in _COUNTRY_ONLY or v in ("—", "-"):
        return ""
    return v


class ExcelWriter:
    def __init__(self, config: dict):
        self.config = config
        self._lock = threading.Lock()
        self._wb: Optional[openpyxl.Workbook] = None
        self._ws: Optional[openpyxl.Worksheet] = None
        self._spec_path: Optional[str] = None
        self._total_rows = 0
        self._save_counter = 0
        self._header_map: Optional[dict] = None
        self._headers: list[str] = []

    # --- Loading ------------------------------------------------------------

    def load_spec(self, path: str) -> tuple[list[str], int]:
        self._wb = openpyxl.load_workbook(path)
        self._ws = self._wb.active
        self._spec_path = path

        raw_headers = [str(h.value) if h.value else "" for h in self._ws[1]]
        self._headers = raw_headers

        data_rows = 0
        for r in range(2, self._ws.max_row + 1):
            if any(self._ws.cell(r, c).value not in (None, "") for c in range(1, self._ws.max_column + 1)):
                data_rows += 1
        self._total_rows = data_rows

        self._header_map = self._find_output_headers()
        mapping = self.detect_columns(raw_headers)
        logger.info("Spec loaded: %d rows; columns: name=%s article=%s brand=%s spec=%s uom=%s qty=%s weight=%s position=%s",
                    data_rows, mapping.get("name"), mapping.get("article"),
                    mapping.get("brand"), mapping.get("spec"), mapping.get("uom"),
                    mapping.get("qty"), mapping.get("weight"), mapping.get("position"))
        return raw_headers, data_rows

    @property
    def headers(self) -> list[str]:
        return self._headers

    @property
    def total_rows(self) -> int:
        return self._total_rows

    @property
    def header_map(self) -> Optional[dict]:
        return self._header_map

    @property
    def spec_path(self) -> Optional[str]:
        return self._spec_path

    @property
    def ws(self):
        return self._ws

    # --- Column Detection ---------------------------------------------------

    def detect_columns(self, headers: list, ws=None) -> dict:
        """Системная классификация колонок (см. src.column_classifier).

        Учитывает заголовки И значения колонок (сэмпл первых 50 строк) —
        корректно обрабатывает «Завод-изготовитель», «Код оборудования…»,
        «Масса единицы (кг)» и прочие реальные заголовки спецификаций.
        """
        ws = ws or self._ws
        rows = []
        if ws is not None:
            for r in range(2, min(ws.max_row + 1, 2 + 50)):
                rows.append([ws.cell(r, c).value for c in range(1, ws.max_column + 1)])
        return classify_columns(headers, value_rows=rows).as_dict()

    def _find_output_headers(self) -> dict:
        ws = self._ws
        if ws is None:
            return {"price": 1, "url": 2, "category": 3}

        header_map = {}
        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(1, col).value or "").lower()
            if "цена" in val:
                header_map["price"] = col
            elif "url" in val and "карточка" not in val:
                header_map["url"] = col
            elif "категор" in val:
                header_map["category"] = col

        last_col = ws.max_column
        if "price" not in header_map:
            header_map["price"] = last_col + 1
            ws.cell(1, header_map["price"], "Цена, RUB")
            last_col += 1
        if "url" not in header_map:
            header_map["url"] = last_col + 1
            ws.cell(1, header_map["url"], "URL")
            last_col += 1
        if "category" not in header_map:
            header_map["category"] = last_col + 1
            ws.cell(1, header_map["category"], "Категория")

        return header_map

    def build_item_name(self, row: int, mapping: dict) -> tuple:
        """Собирает наименование из name-колонок (БЕЗ производителя).

        Завод-изготовитель держится отдельно (SpecItem.brand) — он важен
        для выбора правильного товара агентом, а не как часть поискового
        запроса.
        """
        ws = self._ws
        if ws is None:
            return "", "шт", None

        parts = []
        article = None

        for idx in mapping.get("article", []):
            val = str(ws.cell(row, idx + 1).value or "").strip()
            if val and val not in ("None", ""):
                article = val

        for idx in mapping.get("name", []):
            val = str(ws.cell(row, idx + 1).value or "").strip()
            if val and val not in ("None", ""):
                parts.append(val)

        full_name = " ".join(parts)
        uom = str(ws.cell(row, mapping["uom"] + 1).value or "шт") if mapping.get("uom") is not None else "шт"
        return full_name, uom, article

    # --- Writing Results ----------------------------------------------------

    def write_result(self, row_idx: int, state: dict):
        with self._lock:
            ws = self._ws
            hm = self._header_map
        if ws is None or hm is None:
            logger.warning("write_result: ws or header_map is None")
            return

        price_val = state.get("final_price_rub")
        url_val = state.get("card_url", "")
        cat_key = state.get("primary_cat", "")
        sub_key = state.get("primary_subcat", "")

        cat_ru = _CAT_RU_LABELS.get(cat_key, cat_key)
        if sub_key:
            sub_ru = _SUBCAT_RU_LABELS.get(cat_key, {}).get(sub_key, sub_key)
            cat_val = f"{cat_ru} › {sub_ru}"
        else:
            cat_val = cat_ru

        ws.cell(row_idx, hm["price"], price_val)
        ws.cell(row_idx, hm["url"], url_val)
        ws.cell(row_idx, hm["category"], cat_val)
        self._save_counter += 1

    def flush(self):
        output_dir = self.config.get("paths", {}).get("data_output", "data/output")
        self.save_output_copy(output_dir)

    def save_output_copy(self, output_dir: str) -> str:
        """Сохраняет финальную копию с результатами. Оригинальный файл НЕ трогает."""
        path = self._spec_path
        if not path:
            return ""
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_name = f"{Path(path).stem}_{ts}_priced.xlsx"
        out_path = str(output_path / out_name)

        try:
            with self._lock:
                self._wb.save(out_path)
            logger.info(f"save_output_copy: saved to {out_path}")
        except Exception as e:
            logger.error(f"save_output_copy failed: {e}")

        return out_path

    def get_specs(self) -> list[SpecItem]:
        if self._ws is None or self._headers is None:
            return []
        mapping = self.detect_columns(self._headers)
        qty_col = mapping.get("qty")
        specs = []
        for excel_row in range(2, self._ws.max_row + 1):
            name, uom, article = self.build_item_name(excel_row, mapping)
            if not name or name.strip() in ("", "None", "none"):
                continue
            # строки-заголовки разделов («Отопление», «Вентиляция» и т.п.) —
            # без количества — не являются товарами
            if qty_col is not None:
                qty_val = self._ws.cell(excel_row, qty_col + 1).value
                if qty_val is None or str(qty_val).strip() == "":
                    continue
            brand_raw = _clean_brand(self._concat_cells(excel_row, mapping.get("brand", [])))
            name_raw = self._concat_cells(excel_row, mapping.get("name", []))
            spec_raw = self._concat_cells(excel_row, mapping.get("spec", []))
            specs.append(SpecItem(
                text=name,
                article=article or "",
                brand=brand_raw,
                name_raw=name_raw,
                uom=uom,
                spec=spec_raw,
                headers=self._headers,
            ))
        return specs

    def _concat_cells(self, row: int, indices: list[int]) -> str:
        ws = self._ws
        if ws is None:
            return ""
        parts = []
        for idx in indices:
            val = str(ws.cell(row, idx + 1).value or "").strip()
            if val and val not in ("None", ""):
                parts.append(val)
        return " ".join(parts)






