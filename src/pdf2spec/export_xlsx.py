"""Export classified rows to XLSX with original document column headers.

Produces XLSX "как Hermes" — columns match the source PDF specification.
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

HDR_VK = [
    'Поз.', 'Наименование и техническая характеристика',
    'Тип, марка, обозначение документа, опросного листа', 'Код продукции',
    'Поставщик', 'Ед. измерения', 'Кол.', 'Масса 1 ед., кг', 'Примечание',
]
HDR_OV = [
    'Позиция', 'Наименование и техническая характеристика',
    'Тип, марка, обозначение документа, опросного листа',
    'Код оборудования, изделия, материала', 'Завод-изготовитель',
    'Единица измерения', 'Количество', 'Масса единицы (кг)', 'Примечание',
]

KEYS = ['poz', 'name', 'type', 'code', 'supplier', 'unit', 'qty', 'mass', 'note']

HEADER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
HEADER_FONT = Font(bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def export_xlsx(
    rows: list[dict],
    output_path: str | Path,
    template: str = 'OV',
) -> Path:
    """Export classified rows to XLSX.

    Args:
        rows: list of dicts with role/name/type/code/supplier/unit/qty/mass/note/poz
        output_path: path for the output XLSX file
        template: 'VK' or 'OV' — determines column headers

    Returns:
        Path to the created XLSX file.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    hdr = HDR_OV if template == 'OV' else HDR_VK
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Спецификация'

    for col, h in enumerate(hdr, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical='center')

    for ri, row in enumerate(rows, 2):
        role = row.get('role', 'item')
        if role == 'header':
            values = [row.get('poz', ''), row.get('name', ''), '', '', '', '', '', '', '']
        elif role == 'component':
            values = [
                '', row.get('name', ''), row.get('type', ''), '',
                row.get('supplier', ''), '', '', '', row.get('note', ''),
            ]
        else:
            values = [row.get(k, '') for k in KEYS]

        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='top')

    widths = [8, 50, 30, 15, 20, 12, 8, 12, 25]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    wb.save(output_path)
    return output_path
