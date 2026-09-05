import sys
import yaml
import logging
from pathlib import Path
from datetime import datetime
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                                 QHBoxLayout, QGridLayout, QSplitter, QTableWidget, QTableWidgetItem,
                                   QPushButton, QLabel, QFileDialog, QProgressBar,
                                    QComboBox, QLineEdit, QTextBrowser,
                                    QDialog, QDialogButtonBox, QMessageBox, QMenu,
                                    QStyleFactory, QCheckBox, QHeaderView, QDoubleSpinBox,
                                    QTabWidget, QSizePolicy, QFrame, QLayout, QGroupBox)
from PySide6.QtCore import QObject, Signal, Qt, QTimer, QUrl, QThread, QEvent
from PySide6.QtGui import QDesktopServices, QColor, QPainter

from src.theme import Theme, TOKENS, apply_theme, detect_system_theme
from src import llm_providers
from src import icons as ui_icons
from src.pdf_parser.runner import PdfParserRunner
from src.pdf_parser.review_dialog import ReviewDialog
from src.pdf_parser.feedback import FeedbackCollector
try:
    from src.pdf2spec.runner_v2 import Pdf2SpecRunner
    _HAS_V2 = True
except ImportError:
    _HAS_V2 = False
from src.toast import ToastManager
from src.widget_base import paint_styled_background, setup_shadow
from src.excel_writer import ExcelWriter
from src.mcp_agent_runner import MCPAgentRunner
from gui.graph_assistant import AssistantToolPanel
from gui.graph_explorer import GraphExplorerWidget
from gui.agent_monitor import AgentMonitorPanel
from gui.metrics_panel import MetricsPanel
from gui.spinner_widget import SpinnerWidget
from src.graph_engine import GraphEngine
from src.skip_registry import SkipRegistry


class LogReceiver(QObject):
    log_received = Signal(str, str, str)

_log_receiver = LogReceiver()


class UiLogHandler(logging.Handler):
    def emit(self, record):
        try:
            msg = self.format(record)
            level = {"INFO": "INFO", "WARNING": "WARN", "ERROR": "ERR", "CRITICAL": "ERR"}.get(record.levelname, "INFO")
            phase = getattr(record, 'phase', record.name)
            _log_receiver.log_received.emit(level, phase, msg)
        except Exception:
            pass


logging.basicConfig(level=logging.DEBUG, force=True, format='%(levelname)s:%(name)s:%(message)s')
logging.getLogger().addHandler(UiLogHandler())

_log_dir = Path(__file__).parent / "logs"
_log_dir.mkdir(parents=True, exist_ok=True)
_file_handler = logging.handlers.RotatingFileHandler(
    _log_dir / "runtime.log", encoding="utf-8",
    maxBytes=5*1024*1024, backupCount=3, delay=True
)
_file_handler.setLevel(logging.DEBUG)
_file_handler.setFormatter(logging.Formatter('%(asctime)s [%(levelname)s] [%(name)s] %(message)s'))
logging.getLogger().addHandler(_file_handler)
logging.getLogger("pricer").setLevel(logging.DEBUG)
for noisy in ['websockets', 'asyncio', 'urllib3', 'httpx', 'httpcore', 'httpcore.http11', 'httpcore.connection']:
    logging.getLogger(noisy).setLevel(logging.WARNING)

logger = logging.getLogger(__name__)


class ModelsFetchWorker(QThread):
    fetched = Signal(int, str, list, str)

    def __init__(self, generation, provider_id, base_url, api_key):
        super().__init__()
        self._generation = generation
        self._provider_id = provider_id
        self._base_url = base_url
        self._api_key = api_key

    def run(self):
        try:
            models = llm_providers.get_models_refreshed(
                self._base_url, self._provider_id, self._api_key
            )
            self.fetched.emit(self._generation, self._provider_id, models, "")
        except Exception as e:
            self.fetched.emit(self._generation, self._provider_id, [], str(e))


class SettingsDialog(QDialog):
    def __init__(self, config, parent=None, theme_name=Theme.DARK):
        super().__init__(parent)
        self.setWindowTitle("Настройки LLM")
        self.config = config
        self._theme_name = theme_name
        self._tokens = TOKENS.get(theme_name, TOKENS[Theme.DARK])
        self._workers: set = set()
        self._fetch_generation = 0
        self._pending_models = None
        self._expected_model = ""
        setup_shadow(self, self._tokens)

        lm = config.get("llm", {}) or {}
        root = QVBoxLayout(self)
        root.setSpacing(10)
        root.setContentsMargins(16, 16, 16, 16)

        title = QLabel("Настройки LLM")
        title.setProperty("heading", True)
        root.addWidget(title)

        prov_group = QGroupBox("Провайдер")
        prov_layout = QVBoxLayout(prov_group)
        prov_layout.setSpacing(6)
        self.provider_combo = QComboBox()
        saved_pid = str(lm.get("provider") or "lmstudio")
        for pid, prov in llm_providers.PROVIDERS.items():
            suffix = "" if prov.requires_key else "   ·   локальный"
            self.provider_combo.addItem(f"{prov.name}{suffix}", pid)
        idx = self.provider_combo.findData(saved_pid)
        self.provider_combo.setCurrentIndex(idx if idx >= 0 else 0)
        prov_layout.addWidget(self.provider_combo)
        self.provider_desc = QLabel()
        self.provider_desc.setProperty("muted", True)
        self.provider_desc.setWordWrap(True)
        prov_layout.addWidget(self.provider_desc)
        root.addWidget(prov_group)

        conn_group = QGroupBox("Подключение")
        grid = QGridLayout(conn_group)
        grid.setVerticalSpacing(6)
        grid.setHorizontalSpacing(8)
        grid.setColumnStretch(1, 1)

        lbl_base = QLabel("Base URL:")
        lbl_base.setMinimumWidth(90)
        self.base_url_edit = QLineEdit()
        self.btn_url_default = QPushButton("Сбросить")
        self.btn_url_default.setObjectName("ghost")
        self.btn_url_default.setToolTip("Вернуть Base URL провайдера по умолчанию")
        grid.addWidget(lbl_base, 0, 0)
        grid.addWidget(self.base_url_edit, 0, 1)
        grid.addWidget(self.btn_url_default, 0, 2)

        lbl_key = QLabel("API-ключ:")
        lbl_key.setMinimumWidth(90)
        self.api_key_edit = QLineEdit()
        self.api_key_edit.setEchoMode(QLineEdit.Password)
        self.api_key_edit.setPlaceholderText("пусто → парсится из системы при каждом запуске")
        self.btn_key_system = QPushButton("Из системы")
        self.btn_key_system.setToolTip("Подставить ключ из env / opencode auth.json / hermes .env")
        ui_icons.attach(self.btn_key_system, "key", self._tokens["text-primary"], 16)
        grid.addWidget(lbl_key, 1, 0)
        grid.addWidget(self.api_key_edit, 1, 1)
        grid.addWidget(self.btn_key_system, 1, 2)

        self.key_source_label = QLabel()
        self.key_source_label.setProperty("muted", True)
        self.key_source_label.setWordWrap(True)
        grid.addWidget(self.key_source_label, 2, 1, 1, 2)

        lbl_model = QLabel("Модель:")
        lbl_model.setMinimumWidth(90)
        self.model_combo = QComboBox()
        self.model_combo.setEditable(True)
        self.model_combo.setInsertPolicy(QComboBox.NoInsert)
        self.model_combo.setMaxVisibleItems(20)
        self.model_combo.lineEdit().installEventFilter(self)
        self.model_combo.view().installEventFilter(self)
        self.btn_refresh_models = QPushButton("Обновить")
        self.btn_refresh_models.setToolTip("Получить актуальный список моделей с сервера")
        ui_icons.attach(self.btn_refresh_models, "refresh", self._tokens["text-primary"], 16)
        grid.addWidget(lbl_model, 3, 0)
        grid.addWidget(self.model_combo, 3, 1)
        grid.addWidget(self.btn_refresh_models, 3, 2)
        root.addWidget(conn_group)

        gen_group = QGroupBox("Параметры генерации")
        gen_grid = QGridLayout(gen_group)
        gen_grid.setVerticalSpacing(6)
        gen_grid.setHorizontalSpacing(8)
        lbl_temp = QLabel("Температура:")
        self.temperature_spin = QDoubleSpinBox()
        self.temperature_spin.setRange(0.0, 2.0)
        self.temperature_spin.setSingleStep(0.05)
        self.temperature_spin.setDecimals(2)
        self.temperature_spin.setValue(float(lm.get("temperature", 0.3)))
        lbl_timeout = QLabel("Таймаут (с):")
        self.timeout_spin = QDoubleSpinBox()
        self.timeout_spin.setRange(10, 600)
        self.timeout_spin.setSingleStep(10)
        self.timeout_spin.setDecimals(0)
        self.timeout_spin.setValue(int(lm.get("timeout", 150)))
        gen_grid.addWidget(lbl_temp, 0, 0)
        gen_grid.addWidget(self.temperature_spin, 0, 1)
        gen_grid.addWidget(lbl_timeout, 0, 2)
        gen_grid.addWidget(self.timeout_spin, 0, 3)
        gen_grid.setColumnStretch(1, 1)
        gen_grid.setColumnStretch(3, 1)
        root.addWidget(gen_group)

        pdf_group = QGroupBox("PDF-парсер")
        pdf_layout = QVBoxLayout(pdf_group)
        pdf_row = QHBoxLayout()
        lbl_pipeline = QLabel("Парсер:")
        lbl_pipeline.setMinimumWidth(90)
        self.pipeline_combo = QComboBox()
        self.pipeline_combo.addItem("v2 — методология Hermes (рекомендуется)", "v2")
        self.pipeline_combo.addItem("legacy — старый парсер", "legacy")
        saved_pipeline = config.get("pdf_parser", {}).get("pipeline", "v2")
        idx = self.pipeline_combo.findData(saved_pipeline)
        self.pipeline_combo.setCurrentIndex(idx if idx >= 0 else 0)
        pdf_row.addWidget(lbl_pipeline)
        pdf_row.addWidget(self.pipeline_combo, 1)
        pdf_layout.addLayout(pdf_row)
        pipeline_desc = QLabel(
            "v2: PyMuPDF find_tables + классификация + mother-child + LLM-ревью. "
            "legacy: pdf-inspector/MinerU + regex."
        )
        pipeline_desc.setProperty("muted", True)
        pipeline_desc.setWordWrap(True)
        pdf_layout.addWidget(pipeline_desc)
        root.addWidget(pdf_group)

        test_row = QHBoxLayout()
        self.btn_test = QPushButton("Проверить подключение")
        ui_icons.attach(self.btn_test, "cable", self._tokens["text-primary"], 16)
        self.status_label = QLabel("")
        self.status_label.setWordWrap(True)
        test_row.addWidget(self.btn_test)
        test_row.addWidget(self.status_label, 1)
        root.addLayout(test_row)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        ok_btn = btns.button(QDialogButtonBox.Ok)
        if ok_btn:
            ok_btn.setObjectName("primary")
        btns.accepted.connect(self.save_and_accept)
        btns.rejected.connect(self.reject)
        root.addWidget(btns)

        self.provider_combo.currentIndexChanged.connect(self._on_provider_changed)
        self.btn_url_default.clicked.connect(self._reset_base_url)
        self.btn_key_system.clicked.connect(self._use_system_key)
        self.api_key_edit.textChanged.connect(self._refresh_key_badge)
        self.btn_refresh_models.clicked.connect(self._start_fetch)
        self.btn_test.clicked.connect(self._test_connection)

        self.setMinimumWidth(540)
        self._apply_provider(initial=True)

    def paintEvent(self, event):
        painter = QPainter(self)
        paint_styled_background(self, painter, self._tokens)

    def _current_provider(self):
        return llm_providers.get_provider(self.provider_combo.currentData())

    def _saved_base_url(self, provider_id):
        providers_cfg = ((self.config.get("llm", {}) or {}).get("providers")) or {}
        pcfg = providers_cfg.get(provider_id) or {}
        return (
            (pcfg.get("base_url") or "").strip()
            or llm_providers.resolve_base_url_override(provider_id)
            or llm_providers.get_provider(provider_id).base_url
        )

    def _effective_api_key(self):
        key = self.api_key_edit.text().strip()
        if key:
            return key
        key, _src = llm_providers.resolve_api_key(self.provider_combo.currentData())
        return key

    def _refresh_key_badge(self):
        if not self._current_provider().requires_key:
            self.key_source_label.setText("Ключ не требуется — локальный сервер")
            return
        manual = self.api_key_edit.text().strip()
        if manual:
            source = llm_providers.SOURCE_OVERRIDE
            fingerprint = llm_providers.key_fingerprint(manual)
        else:
            key, source = llm_providers.resolve_api_key(self.provider_combo.currentData())
            fingerprint = llm_providers.key_fingerprint(key)
            if not key:
                self.key_source_label.setText(
                    "⚠ Ключ не найден в системе (env / opencode auth.json / hermes .env)"
                )
                return
        self.key_source_label.setText(f"Ключ: {source} · {fingerprint}")

    def _apply_provider(self, initial=False):
        prov = self._current_provider()
        self.provider_desc.setText(prov.description)
        self.base_url_edit.setText(self._saved_base_url(prov.id))
        if not initial:
            self.api_key_edit.clear()
        model = prov.default_model
        if initial:
            configured = str((self.config.get("llm", {}) or {}).get("model") or "").strip()
            cached = llm_providers.cached_models(prov.id) or []
            known_ids = {m.get("id") for m in cached}
            if configured and configured != "local-model" and (
                not known_ids or configured in known_ids
            ):
                model = configured
        self._expected_model = model
        self.model_combo.setCurrentText(model)
        self._refresh_key_badge()
        cached = llm_providers.cached_models(prov.id)
        if cached is not None and not self.model_combo.view().isVisible():
            self._populate_models(cached, select=model)
            self._reconcile_expected_model(cached)
        self._start_fetch()

    def _reconcile_expected_model(self, models):
        """Модель из конфига чужому провайдеру не достаётся: если ожидаемый
        выбор отсутствует в актуальном списке — откат к дефолту провайдера."""
        ids = {m.get("id") for m in models}
        current = (self.model_combo.currentData() or self.model_combo.currentText() or "").strip()
        if current and ids and current not in ids:
            fallback = self._expected_model or self._current_provider().default_model
            self.model_combo.setCurrentText(fallback)

    def _event_target_is_model_combo(self, obj):
        line_edit = self.model_combo.lineEdit()
        return obj is line_edit or obj is self.model_combo.view()

    def eventFilter(self, obj, event):
        # Клик по полю editable-комбобокса открывает попап (как у не-editable).
        if self._event_target_is_model_combo(obj):
            if event.type() == QEvent.MouseButtonPress:
                if obj is self.model_combo.lineEdit() and not self.model_combo.view().isVisible():
                    self.model_combo.showPopup()
                    return False
            elif event.type() == QEvent.Hide and self._pending_models is not None:
                models, select = self._pending_models
                self._pending_models = None
                self._populate_models(models, select=select)
        return super().eventFilter(obj, event)

    def _on_provider_changed(self):
        self._apply_provider(initial=False)

    def _reset_base_url(self):
        pid = self.provider_combo.currentData()
        override = llm_providers.resolve_base_url_override(pid)
        self.base_url_edit.setText(override or self._current_provider().base_url)

    def _use_system_key(self):
        pid = self.provider_combo.currentData()
        llm_providers.set_manual_key(pid, "")
        self.api_key_edit.clear()
        key, src = llm_providers.resolve_api_key(pid)
        if not key:
            self._set_status("Ключ не найден ни в env, ни в opencode auth.json, ни в hermes .env", danger=True)
            self._refresh_key_badge()
            return
        self.api_key_edit.setText(key)
        self._set_status(f"Ключ подставлен из системы ({src})", success=True)
        self._refresh_key_badge()

    def _populate_models(self, models, select=""):
        # Репопуляция при открытом попапе заставляет Qt его закрыть — откладываем.
        if self.model_combo.view().isVisible():
            self._pending_models = ([dict(m) for m in models], select)
            return
        current = select or (self.model_combo.currentData() or self.model_combo.currentText() or "").strip()
        self.model_combo.clear()
        for m in models:
            mid = m.get("id", "")
            name = m.get("name") or ""
            label = mid if not name else f"{mid}   ·   {name}"
            self.model_combo.addItem(label, mid)
        if current:
            i = self.model_combo.findData(current)
            if i >= 0:
                self.model_combo.setCurrentIndex(i)
            else:
                self.model_combo.setCurrentText(current)

    def _start_fetch(self):
        """Живой запрос /models при каждом открытии/переключении (как в opencode/hermes).

        Кэш используется только для мгновенной подстановки до ответа и как
        fallback при ошибке сети. Ответы устаревших поколений отбрасываются.
        """
        pid = self.provider_combo.currentData()
        base = self.base_url_edit.text().strip()
        if not base:
            self._set_status("Base URL не задан", danger=True)
            return
        self._fetch_generation += 1
        self.btn_refresh_models.setEnabled(False)
        self.btn_test.setEnabled(False)
        self._set_status("Обновление списка моделей…")
        worker = ModelsFetchWorker(
            self._fetch_generation, pid, base, self._effective_api_key()
        )
        worker.fetched.connect(self._on_models_fetched)
        worker.finished.connect(lambda w=worker: self._discard_worker(w))
        self._workers.add(worker)
        worker.start()

    def _discard_worker(self, worker):
        self._workers.discard(worker)
        worker.deleteLater()

    def _test_connection(self):
        self._start_fetch()

    def _on_models_fetched(self, generation, provider_id, models, error):
        if generation != self._fetch_generation or provider_id != self.provider_combo.currentData():
            return
        self.btn_refresh_models.setEnabled(True)
        self.btn_test.setEnabled(True)
        if error:
            cached = llm_providers.cached_models(provider_id)
            if cached:
                self._populate_models(cached)
                self._reconcile_expected_model(cached)
                self._set_status(f"Сервер недоступен ({error}); показан сохранённый список", danger=True)
            else:
                self._set_status(f"Ошибка: {error}", danger=True)
            return
        self._populate_models(models)
        self._reconcile_expected_model(models)
        count = self.model_combo.count()
        self._set_status(f"Подключение OK · {count} моделей", success=True)

    def _set_status(self, text, success=False, danger=False):
        t = self._tokens
        color = t["success"] if success else t["danger"] if danger else t["text-muted"]
        self.status_label.setText(text)
        self.status_label.setStyleSheet(f"color:{color};")

    def _current_model_id(self) -> str:
        """id модели из editable-комбобокса.

        ВАЖНО: currentData() в editable-комбобоксе хранит данные ТЕКУЩЕГО
        ИНДЕКСА, а не введённого текста. Если пользователь ввёл модель вручную
        (или пришёл асинхронный populate), currentData() может отдавать СТАРУЮ
        модель — тогда в settings.yaml сохранялся бы не выбор пользователя,
        а устаревший индекс (very-high вместо very-low). Надёжный источник —
        отображаемый текст: либо 'id · имя' (выбран пункт), либо введённый id.
        """
        return llm_providers.model_id_from_combo_text(self.model_combo.currentText())

    def save_and_accept(self):
        prov = self._current_provider()
        model_id = self._current_model_id()
        if not model_id:
            # Запасной вариант, если текст пуст (например, чистый data).
            model_id = (self.model_combo.currentData() or "").strip()
        if not model_id:
            self._set_status("Укажите модель", danger=True)
            return
        base_url = self.base_url_edit.text().strip() or prov.base_url
        typed_key = self.api_key_edit.text().strip()
        system_key, _src = llm_providers.resolve_api_key(prov)
        # Ручной ключ фиксируем только если он реально отличается от системного,
        # иначе диагностика источника ключа вводит в заблуждение.
        llm_providers.set_manual_key(
            prov.id, typed_key if typed_key and typed_key != system_key else ""
        )
        from src.config_loader import save_llm_settings
        save_llm_settings(
            provider=prov.id,
            model=model_id,
            temperature=float(self.temperature_spin.value()),
            timeout=int(self.timeout_spin.value()),
            base_urls={prov.id: base_url},
        )

        pipeline = self.pipeline_combo.currentData()
        if pipeline:
            self.config.setdefault("pdf_parser", {})["pipeline"] = pipeline
            from src.config_loader import save_pdf_parser_settings
            save_pdf_parser_settings(self.config.get("pdf_parser", {}))

        logger.info("LLM settings saved: provider=%s model=%s base=%s pipeline=%s",
                     prov.id, model_id, base_url, pipeline)
        self.accept()

    def _wait_workers(self):
        for worker in list(self._workers):
            if worker.isRunning():
                worker.wait(5000)
        self._workers.clear()

    def reject(self):
        self._wait_workers()
        super().reject()

    def closeEvent(self, event):
        self._wait_workers()
        super().closeEvent(event)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Pricer Vision v31.0")
        self.resize(1600, 950)
        ui_icons.register()
        self.config = self._load_config()
        self.excel_writer = ExcelWriter(self.config)
        self._engine = GraphEngine("data/pricer.db")
        self._engine.build()
        from pathlib import Path
        yaml_path = "config/categories_and_sites.yaml"
        if Path(yaml_path).exists():
            self._engine.load_yaml_seed(yaml_path)

        self._processing_active = False
        self._spec_path = None
        self._total_rows = 0
        self._skip_registry = SkipRegistry()
        self._skip_reconciling = False
        self._restored_results = []
        self._restored_row_indices = set()
        self._retry_row = None
        self._restored_caches = None
        self._restored_audit_id = ""
        self._original_restored_results = []
        self._session_log_entries = []
        from src.approach_relevance import load_rules
        load_rules()
        from src.config_loader import load_settings
        self._current_theme = load_settings().get("ui", {}).get("theme") or detect_system_theme()
        self._spinner_color = "#89b4fa"
        self._spinner_timer = QTimer(self)
        self._spinner_timer.setInterval(120)
        self._spinner_timer.timeout.connect(self._tick_spinner)
        self._spinner = SpinnerWidget(size=14, color=self._spinner_color, spacing=0.5)

        self._setup_ui()
        self._connect_signals()

        apply_theme(QApplication.instance(), self._current_theme)
        self.toast_manager = ToastManager(self)

        self._session_check_scheduled = False

    def showEvent(self, event):
        super().showEvent(event)
        # Проверка сохранённой сессии — только ПОСЛЕ того, как окно показано
        # и event loop запущен. Раньше вызов шёл из __init__ через QTimer.singleShot(100)
        # ДО showMaximized()/app.exec() — модальный dlg.exec() открывал вложенный
        # event loop в неотрисованном окне, и последующие load_spec/QMessageBox
        # зависали (UI «не отвечает»).
        if not self._session_check_scheduled:
            self._session_check_scheduled = True
            QTimer.singleShot(200, self._safe_check_last_session)

    def _safe_check_last_session(self):
        """Запуск проверки сессии с защитой: ошибка не должна вешать UI."""
        try:
            self._check_last_session()
        except Exception as e:
            logger.error("Session restore check failed: %s", e, exc_info=True)

    def closeEvent(self, event):
        if self._processing_active and hasattr(self, '_runner') and self._runner:
            self._runner.stop()
            self._runner.wait(3000)
        self._auto_save_session()
        if hasattr(self, 'graph_widget'):
            self.graph_widget._physics.stop()
        super().closeEvent(event)

    def _auto_save_session(self):
        """Сохраняет текущую сессию в data/sessions/_current.json."""
        from src.session_manager import save_session, auto_save_path
        if not self._spec_path:
            return
        state = self._build_session_state()
        if state["results"] or state["negative_cache"] or state["skip_registry"].get("marked"):
            try:
                save_session(auto_save_path(), state)
                logger.info("Auto-saved session to %s", auto_save_path())
            except Exception as e:
                logger.error("Auto-save session failed: %s", e)

    def _build_session_state(self) -> dict:
        """Собирает текущее состояние сессии для сохранения."""
        runner = getattr(self, '_runner', None)
        caches = {}
        if runner and hasattr(runner, '_restored_caches') and runner._restored_caches:
            caches = runner._restored_caches
        elif self._restored_caches:
            caches = self._restored_caches
        audit_id = ""
        if runner and hasattr(runner, 'audit_session_id'):
            audit_id = runner.audit_session_id
        elif self._restored_audit_id:
            audit_id = self._restored_audit_id
        return {
            "spec_path": self._spec_path or "",
            "total_rows": self._total_rows,
            "results": self._restored_results,
            "negative_cache": caches.get("negative_cache", {}),
            "site_blacklist": caches.get("site_blacklist", {}),
            "session_facts": caches.get("session_facts", {}),
            "skip_registry": self._skip_registry.to_dict(),
            "run_flags": {
                "reuse_price": self.reuse_price_cb.isChecked(),
                "use_approaches": self.use_approaches_cb.isChecked(),
                "use_site_ranking": self.use_site_ranking_cb.isChecked(),
                "ductwork_enabled": self.ductwork_cb.isChecked(),
            },
            "metrics": {},
            "log_entries": self._session_log_entries[-200:],
            "audit_session_id": audit_id,
        }

    def _check_last_session(self):
        """При запуске: показать диалог выбора сессии, если есть сохранённые."""
        from src.session_manager import (
            list_sessions, has_current_session, load_current_session, auto_save_path,
            recover_corrupted_sessions,
        )
        # Восстанавливаем повреждённые файлы сессий (если есть)
        recover_corrupted_sessions()
        sessions = list_sessions()
        current_has = has_current_session()
        if not sessions and not current_has:
            return
        if current_has:
            # Автосохранённая сессия (_current.json) не попадает в list_sessions —
            # добавляем её в начало списка, чтобы её можно было восстановить.
            cur = load_current_session()
            if cur and (cur.get("results") or cur.get("negative_cache")):
                sessions.insert(0, {
                    "path": auto_save_path(),
                    "saved_at": cur.get("saved_at", ""),
                    "spec_name": cur.get("spec_name", "Текущая (автосохранённая)"),
                    "total_rows": cur.get("total_rows", 0),
                    "processed_count": cur.get("processed_count", 0),
                    "found_count": cur.get("found_count", 0),
                })
            elif current_has and not cur:
                # _current.json повреждён — сообщаем пользователю
                from PySide6.QtWidgets import QMessageBox
                QMessageBox.warning(
                    self, "Сессия повреждена",
                    "Файл автосохранения (_current.json) повреждён.\n"
                    "Предыдущая сессия не может быть восстановлена.\n\n"
                    "Начните новую сессию или загрузите сохранённую из меню «Сессия».")
        if not sessions:
            return
        from gui.session_dialog import SessionDialog
        dlg = SessionDialog(sessions, self)
        if dlg.exec():
            if dlg.selected_session:
                try:
                    self._restore_session(dlg.selected_session)
                except Exception as e:
                    logger.error("Session restore failed: %s", e, exc_info=True)
                    self.add_log("ERR", "session", f"Ошибка восстановления сессии: {e}")

    def _open_session_dialog(self):
        """Открыть диалог выбора сессии по кнопке «Сессия»."""
        from src.session_manager import (
            list_sessions, save_session, auto_save_path,
            has_current_session, load_current_session,
        )
        if self._spec_path and self._restored_results:
            state = self._build_session_state()
            try:
                save_session(auto_save_path(), state)
            except Exception:
                pass
        sessions = list_sessions()
        if has_current_session():
            cur = load_current_session()
            if cur and (cur.get("results") or cur.get("negative_cache")):
                sessions.insert(0, {
                    "path": auto_save_path(),
                    "saved_at": cur.get("saved_at", ""),
                    "spec_name": cur.get("spec_name", "Текущая (автосохранённая)"),
                    "total_rows": cur.get("total_rows", 0),
                    "processed_count": cur.get("processed_count", 0),
                    "found_count": cur.get("found_count", 0),
                })
        dlg = SessionDialog(sessions, self)
        if dlg.exec():
            if dlg.selected_session:
                try:
                    self._restore_session(dlg.selected_session)
                except Exception as e:
                    logger.error("Session restore failed: %s", e, exc_info=True)
                    self.add_log("ERR", "session", f"Ошибка восстановления сессии: {e}")

    @staticmethod
    def _resolve_spec_path(original: str, spec_name: str) -> str | None:
        """Ищет файл спецификации, если оригинальный путь недоступен.

        Порядок поиска:
        1. data/output/ — ищем по имени файла (стем)
        2. data/output/ — pdf_spec_*.xlsx (PDF-конверсия)
        3. Рядом с оригинальным путём (parent dir) —同名 файл
        """
        from pathlib import Path
        stem = Path(original).stem if original else ""
        out_dir = Path("data/output")

        # 1. Точное совпадение по имени в data/output/
        if stem and out_dir.exists():
            for f in out_dir.glob(f"{stem}*.xlsx"):
                return str(f)

        # 2. Последний pdf_spec_*.xlsx в data/output/ (если сессия из PDF)
        if out_dir.exists():
            pdf_specs = sorted(out_dir.glob("pdf_spec_*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
            if pdf_specs:
                return str(pdf_specs[0])

        # 3. Рядом с оригинальным путём (тот же parent, другое имя)
        if original:
            parent = Path(original).parent
            if parent.exists() and stem:
                for f in parent.glob(f"{stem}*.xlsx"):
                    return str(f)

        return None

    def _restore_session(self, session_path: str):
        """Восстанавливает сессию из JSON-файла."""
        from src.session_manager import load_session
        state = load_session(session_path)
        if not state:
            QMessageBox.warning(self, "Ошибка", "Не удалось загрузить сессию.")
            return

        spec_path = state.get("spec_path", "")
        if not spec_path or not Path(spec_path).exists():
            # Файл спецификации перемещён/удалён — ищем по имени в data/output/
            # и рядом с оригинальным путём.
            spec_path = self._resolve_spec_path(spec_path, state.get("spec_name", ""))
            if not spec_path:
                QMessageBox.warning(self, "Ошибка",
                                    f"Файл спецификации не найден:\n{state.get('spec_path', '?')}\n\n"
                                    "Проверьте, что файл не был перемещён или удалён.")
                return

        self.load_spec(path=spec_path)
        if not self._spec_path:
            return

        self._skip_registry.from_dict(state.get("skip_registry", {}))

        self._restored_results = state.get("results", [])
        # Дедупликация: артефакт предыдущих багов — в файле могли оказаться
        # дублирующиеся записи. Оставляем первый экземпляр каждого spec_text
        # (с ценой приоритетнее, но порядок в JSON и так ставит их раньше).
        _seen_specs = set()
        _deduped = []
        for r in self._restored_results:
            st = r.get("spec_text", "")
            if st and st in _seen_specs:
                continue
            _seen_specs.add(st)
            _deduped.append(r)
        self._restored_results = _deduped
        self._original_restored_results = list(self._restored_results)
        self._restored_row_indices = set()
        for result in self._restored_results:
            excel_row = result.get("excel_row", 0)
            if excel_row >= 2:
                self._restored_row_indices.add(excel_row - 2)

        flags = state.get("run_flags", {})
        self.reuse_price_cb.setChecked(flags.get("reuse_price", True))
        self.use_approaches_cb.setChecked(flags.get("use_approaches", True))
        self.use_site_ranking_cb.setChecked(flags.get("use_site_ranking", True))
        self.ductwork_cb.setChecked(flags.get("ductwork_enabled", False))

        self._restored_caches = {
            "negative_cache": state.get("negative_cache", {}),
            "site_blacklist": state.get("site_blacklist", {}),
            "session_facts": state.get("session_facts", {}),
        }

        self._restored_audit_id = state.get("audit_session_id", "")

        self._session_log_entries = state.get("log_entries", [])
        for entry in self._session_log_entries:
            self.add_log(entry.get("level", "INFO"), entry.get("phase", "session"),
                         entry.get("msg", ""))

        # Заполнение таблицы результатов ОТЛОЖЕННО (по одной строке на итерацию
        # event loop). При старте окно ещё в процессе раскладки — синхронный
        # _on_row_done → results_table.scrollToBottom() в QTimer-колбэке вызывает
        # re-entrancy deadlock (UI «не отвечает»).
        _pending = [r for r in self._restored_results if (r.get("excel_row") or 0) >= 2]

        def _populate():
            if not _pending:
                self._finish_restore()
                return
            result = _pending.pop(0)
            try:
                result["restored"] = True
                self._on_row_done((result.get("excel_row") or 0) - 2, result)
            except Exception as e:
                logger.error("Session restore row failed: %s", e, exc_info=True)
            QTimer.singleShot(0, _populate)

        QTimer.singleShot(0, _populate)

    def _finish_restore(self):
        self.add_log("INFO", "session",
                     f"Сессия восстановлена: {len(self._restored_results)} результатов, "
                     f"{len(self._restored_row_indices)} строк обработано")
        try:
            self.toast_manager.success(f"Сессия восстановлена ({len(self._restored_results)} результатов)")
        except Exception:
            pass

    def _load_config(self):
        config_path = Path(__file__).parent / "config" / "settings.yaml"
        with open(config_path, "r", encoding="utf-8") as f:
            return yaml.safe_load(f) or {}

    def _setup_ui(self):
        cw = QWidget()
        self.setCentralWidget(cw)
        main_layout = QVBoxLayout(cw)
        main_layout.setContentsMargins(10, 2, 10, 10)
        main_layout.setSpacing(4)

        self._icon_specs: list = []

        def _icon(widget, key, px=18):
            _tt = TOKENS.get(self._current_theme, TOKENS[Theme.DARK])
            ui_icons.attach(widget, key, _tt["text-primary"], px)
            self._icon_specs.append((widget, key, px))
            return widget

        btn_frame = QFrame()
        btn_frame.setFrameShape(QFrame.NoFrame)
        btn_frame.setLayout(QHBoxLayout()); btn_frame.layout().setContentsMargins(6, 3, 6, 3); btn_frame.layout().setSpacing(10)
        btn_frame.setFixedHeight(38)
        top_bar = btn_frame.layout()

        self.load_btn = _icon(QPushButton("Загрузить Excel"), "bar_chart", 18)
        self.load_btn.clicked.connect(self.load_spec)
        top_bar.addWidget(self.load_btn)

        self.pdf_btn = _icon(QPushButton("Загрузить PDF"), "description", 18)
        self.pdf_btn.clicked.connect(self._load_pdf)
        top_bar.addWidget(self.pdf_btn)

        self.start_btn = QPushButton("Старт")
        self.start_btn.setObjectName("primary")
        self.start_btn.clicked.connect(self.start_processing)
        self.start_btn.setEnabled(False)
        top_bar.addWidget(self.start_btn)

        self.stop_btn = QPushButton("Стоп")
        self.stop_btn.setObjectName("danger")
        self.stop_btn.clicked.connect(self.stop_processing)
        self.stop_btn.setEnabled(False)
        top_bar.addWidget(self.stop_btn)

        self.study_btn = _icon(QPushButton("Обучение"), "menu_book", 18)
        self.study_btn.clicked.connect(self._open_study_tool)
        top_bar.addWidget(self.study_btn)

        self.clear_skip_btn = QPushButton("Снять отметки")
        self.clear_skip_btn.setToolTip("Снять все отметки «пропустить» в предпросмотре")
        self.clear_skip_btn.setEnabled(False)
        self.clear_skip_btn.clicked.connect(self._clear_skip_marks)
        top_bar.addWidget(self.clear_skip_btn)

        self.session_btn = _icon(QPushButton("Сессия"), "save", 18)
        self.session_btn.setToolTip("Сохранить/загрузить сессию обработки")
        self.session_btn.clicked.connect(self._open_session_dialog)
        top_bar.addWidget(self.session_btn)

        from src.config_loader import load_settings
        self.headless_cb = _icon(QCheckBox("Headless"), "visibility_off", 15)
        self.headless_cb.setChecked(load_settings().get("browser", {}).get("headless", True))
        self.headless_cb.toggled.connect(self._on_headless_toggle)
        top_bar.addWidget(self.headless_cb)

        self.backend_combo = QComboBox()
        self.backend_combo.setToolTip("Браузерный движок для агента: антидетект Camoufox, Playwright MCP или Nodriver")
        self.backend_combo.addItem("Camoufox (антидетект)", "camoufox")
        self.backend_combo.addItem("Playwright MCP", "playwright")
        self.backend_combo.addItem("Nodriver", "nodriver")
        current_backend = load_settings().get("browser", {}).get("backend", "camoufox")
        bi = self.backend_combo.findData(current_backend)
        self.backend_combo.setCurrentIndex(bi if bi >= 0 else 0)
        self.backend_combo.currentIndexChanged.connect(self._on_backend_change)
        top_bar.addWidget(self.backend_combo)

        top_bar.addStretch()

        self.settings_btn = QPushButton("Настройки")
        self.settings_btn.clicked.connect(self.open_settings)
        top_bar.addWidget(self.settings_btn)

        self.deps_btn = _icon(QPushButton("Зависимости"), "extension", 18)
        self.deps_btn.setToolTip("Обновить и изменить версии зависимостей (pip, @playwright/mcp)")
        self.deps_btn.clicked.connect(self.open_dependency_manager)
        top_bar.addWidget(self.deps_btn)

        self.rules_btn = _icon(QPushButton("Правила сопоставления"), "rule", 18)
        self.rules_btn.setToolTip("Настроить правила сопоставления наименований товаров (без правки кода)")
        self.rules_btn.clicked.connect(self.open_rules_editor)
        top_bar.addWidget(self.rules_btn)

        self.theme_btn = QPushButton("Тема")
        self.theme_btn.clicked.connect(self._toggle_theme)
        top_bar.addWidget(self.theme_btn)

        main_layout.addWidget(btn_frame, 0)

        # Панель «Режим поиска»: три независимых флажка памяти агента.
        run_frame = QFrame()
        run_frame.setFrameShape(QFrame.NoFrame)
        run_layout = QVBoxLayout(run_frame)
        run_layout.setContentsMargins(8, 2, 8, 2)
        run_layout.setSpacing(2)
        run_layout.setAlignment(Qt.AlignLeft)
        run_title = QLabel("РЕЖИМ ПОИСКА")
        run_title.setProperty("muted", True)
        run_layout.addWidget(run_title)
        run_checks = QHBoxLayout()
        run_checks.setSpacing(14)
        from src.config_loader import get_run_flags
        _run_flags = get_run_flags()
        self.reuse_price_cb = QCheckBox("Цены из памяти")
        self.reuse_price_cb.setChecked(_run_flags["reuse_price"])
        self.reuse_price_cb.setToolTip("Переиспользовать ранее подтверждённые цены (rule-8 и кэш). Снимите для чистого поиска")
        self.reuse_price_cb.toggled.connect(self._on_run_mode_toggle)
        run_checks.addWidget(self.reuse_price_cb)
        self.use_approaches_cb = QCheckBox("Подходы")
        self.use_approaches_cb.setChecked(_run_flags["use_approaches"])
        self.use_approaches_cb.setToolTip("Использовать сохранённые шаги поиска по сайтам (граф-память)")
        self.use_approaches_cb.toggled.connect(self._on_run_mode_toggle)
        run_checks.addWidget(self.use_approaches_cb)
        self.use_site_ranking_cb = QCheckBox("Рейтинг сайтов")
        self.use_site_ranking_cb.setChecked(_run_flags["use_site_ranking"])
        self.use_site_ranking_cb.setToolTip("Начинать поиск с сайтов, где этот тип товара находился быстрее всего")
        self.use_site_ranking_cb.toggled.connect(self._on_run_mode_toggle)
        run_checks.addWidget(self.use_site_ranking_cb)
        from src.config_loader import get_ductwork_enabled
        self.ductwork_cb = QCheckBox("Расчёт воздуховодов")
        self.ductwork_cb.setChecked(get_ductwork_enabled())
        self.ductwork_cb.setToolTip("Воздуховоды и фасонные части считать локально (без обращения к сайтам)")
        self.ductwork_cb.toggled.connect(self._on_run_mode_toggle)
        run_checks.addWidget(self.ductwork_cb)
        run_checks.addSpacing(8)
        run_hint = QLabel("ⓘ Чистый поиск: снять все три флажка")
        run_hint.setProperty("muted", True)
        run_checks.addWidget(run_hint)
        run_checks.addStretch()
        run_layout.addLayout(run_checks)
        run_frame.setFixedHeight(52)
        main_layout.addWidget(run_frame, 0)

        fb_frame = QFrame()
        fb_frame.setFrameShape(QFrame.NoFrame)
        fb_layout = QHBoxLayout(fb_frame)
        fb_layout.setContentsMargins(6, 2, 6, 2)
        fb_layout.setSpacing(10)
        self._spinner.setFixedSize(16, 16)
        fb_layout.addWidget(self._spinner)
        self.status_label = QLabel("Готов")
        self.status_label.setFixedHeight(24)
        fb_layout.addWidget(self.status_label, 1)
        fb_frame.setFixedHeight(28)
        main_layout.addWidget(fb_frame, 0)

        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedHeight(21)
        main_layout.addWidget(self.progress_bar, 0)

        splitter = QSplitter(Qt.Horizontal)
        splitter.setHandleWidth(3)
        self._splitter = splitter

        center = QWidget()
        center_layout = QVBoxLayout(center)
        center_layout.setContentsMargins(0, 0, 0, 0)
        center_layout.setSpacing(0)

        center_tabs = QTabWidget()
        center_tabs.setTabPosition(QTabWidget.South)
        self._center_tabs = center_tabs

        self.results_table = QTableWidget(0, 8)
        self.results_table.setHorizontalHeaderLabels([
            "#", "Спецификация", "Тип", "Цена", "Уверенность", "Время", "Сайт", "URL"
        ])
        self.results_table.setColumnWidth(0, 30)
        self.results_table.setColumnWidth(1, 400)
        self.results_table.setColumnWidth(2, 80)
        self.results_table.setColumnWidth(3, 90)
        self.results_table.setColumnWidth(4, 80)
        self.results_table.setColumnWidth(5, 65)
        self.results_table.setColumnWidth(6, 110)
        self.results_table.horizontalHeader().setSectionResizeMode(7, QHeaderView.Stretch)
        self.results_table.setWordWrap(True)
        self.results_table.verticalHeader().setDefaultSectionSize(30)
        self.results_table.setAlternatingRowColors(True)
        self.results_table.itemDoubleClicked.connect(self._on_url_double_click)
        self.results_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.results_table.customContextMenuRequested.connect(self._on_results_context_menu)
        center_tabs.addTab(self.results_table, "Результаты")

        self.preview_table = QTableWidget(0, 7)
        self.preview_table.setHorizontalHeaderLabels(
            ["Пропустить", "#", "Наименование", "Производитель", "Тип/обозначение", "Артикул", "Кол-во"]
        )
        self.preview_table.setColumnWidth(0, 70)
        self.preview_table.setColumnWidth(1, 30)
        self.preview_table.setColumnWidth(2, 260)
        self.preview_table.setColumnWidth(3, 110)
        self.preview_table.setColumnWidth(4, 140)
        self.preview_table.setColumnWidth(5, 130)
        self.preview_table.setColumnWidth(6, 60)
        self.preview_table.setAlternatingRowColors(True)
        center_tabs.addTab(self.preview_table, "Предпросмотр")

        self.log_browser = QTextBrowser()
        center_tabs.addTab(self.log_browser, "Логи")
        center_tabs.setCurrentIndex(1)  # start on Предпросмотр

        center_layout.addWidget(center_tabs, 2)
        splitter.addWidget(center)

        right_tabs = QTabWidget()
        self._right_tabs = right_tabs
        right_tabs.setTabPosition(QTabWidget.South)
        right_tabs.setMinimumWidth(0)
        right_tabs.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Ignored)

        self.graph_widget = GraphExplorerWidget()
        self.graph_widget.load_graph(self._engine)
        right_tabs.addTab(self.graph_widget, "Граф")

        self.assistant_panel = AssistantToolPanel()
        self.assistant_panel.engine = self._engine
        self.assistant_panel.llm_config = self.config.get("llm", {})
        right_tabs.addTab(self.assistant_panel, "Ассистент")

        monitor_widget = QWidget()
        monitor_layout = QVBoxLayout(monitor_widget)
        monitor_layout.setContentsMargins(0, 0, 0, 0)
        monitor_layout.setSpacing(4)
        self.monitor_panel = AgentMonitorPanel()
        monitor_layout.addWidget(self.monitor_panel, 2)
        self.metrics_panel = MetricsPanel()
        # Плитки метрик фиксированной высоты — не растягиваются, растёт монитор.
        monitor_layout.addWidget(self.metrics_panel, 0, Qt.AlignTop)
        right_tabs.addTab(monitor_widget, "Мониторинг")
        # По умолчанию правая панель открывается на «Мониторинге»
        right_tabs.setCurrentIndex(right_tabs.count() - 1)

        splitter.addWidget(right_tabs)

        splitter.setSizes([700, 500])
        main_layout.addWidget(splitter, 1)

        self._log_data = []
        self._log_mode = "all"

    def _connect_signals(self):
        _log_receiver.log_received.connect(self.add_log)
        self.preview_table.itemChanged.connect(self._on_preview_item_changed)

    def _on_url_double_click(self, item):
        # URL — колонка 7 (не 6 «Сайт»). Открываем ПОЛНЫЙ URL из UserRole
        # (усечённый text() давал битую ссылку → 404).
        if item.column() == 7:
            url = item.data(Qt.UserRole) or item.text()
            if url and url.startswith("http"):
                QDesktopServices.openUrl(QUrl(url))

    def _on_results_context_menu(self, pos):
        """Контекстное меню строки результатов — полный контроль над любой строкой.

        Доступно всегда (независимо от того, найдена цена или нет):
          - Повторить поиск (fresh, с очисткой памяти строки)
          - Удалить результат (полная очистка: таблица + граф + кэши + Excel)
          - Обучить агента на этой позиции
          - Исправить тип товара (переклассификация + перепоиск)
        """
        item = self.results_table.itemAt(pos)
        if not item:
            return
        row = item.row()
        spec_item = self.results_table.item(row, 1)
        if not spec_item:
            return
        spec_text = spec_item.text()
        result = None
        for r in self._restored_results:
            if r.get("spec_text") == spec_text:
                result = r
                break
        if not result:
            # Строка может не иметь result (не обработана) — всё равно даём повтор.
            result = {"spec_text": spec_text, "excel_row": row + 2}

        _rt = TOKENS.get(self._current_theme, TOKENS[Theme.DARK])
        menu = QMenu(self)
        can_edit = not self._processing_active

        action_retry = menu.addAction("Повторить поиск (с нуля)")
        action_retry.setIcon(ui_icons.icon("refresh", _rt["text-primary"], 16))
        action_retry.setEnabled(can_edit)
        action_retry.triggered.connect(lambda: self._confirm_and_retry_row(row, spec_text))

        action_delete = menu.addAction("Удалить результат")
        action_delete.setIcon(ui_icons.icon("delete", _rt["danger"], 16))
        action_delete.setEnabled(can_edit)
        action_delete.triggered.connect(lambda: self._delete_row_result(row, spec_text))

        action_study = menu.addAction("Обучить агента на этой позиции")
        action_study.setIcon(ui_icons.icon("smart_toy", _rt["text-primary"], 16))
        pt = result.get("product_type", "")
        action_study.triggered.connect(
            lambda: self._open_study(spec_text, pt or self._engine.classify_product_type(spec_text)))

        action_type = menu.addAction("Исправить тип товара…")
        action_type.setIcon(ui_icons.icon("edit_note", _rt["text-primary"], 16))
        action_type.setEnabled(can_edit)
        action_type.triggered.connect(lambda: self._fix_row_type(row, spec_text))

        menu.exec(self.results_table.viewport().mapToGlobal(pos))

    def _confirm_and_retry_row(self, table_row: int, spec_text: str):
        """Повтор строки с нуля: подтверждение, очистка памяти, fresh-прогон."""
        if self._processing_active:
            return
        ret = QMessageBox.question(
            self, "Повторить поиск",
            "Повторный поиск очистит сохранённый результат этой строки\n"
            "(цену, URL и память графа/кэша) и запустит поиск с нуля.\n\nПродолжить?",
            QMessageBox.Yes | QMessageBox.No)
        if ret != QMessageBox.Yes:
            return
        self._purge_row_memory(spec_text)
        self._retry_single_row(table_row)

    def _delete_row_result(self, table_row: int, spec_text: str):
        """Полное удаление результата строки: таблица, граф, кэши, Excel."""
        if self._processing_active:
            return
        url, site, price, pt, excel_row = "", "", None, "", None
        for r in self._restored_results:
            if r.get("spec_text") == spec_text:
                url = r.get("url", "")
                site = r.get("site", "")
                price = r.get("price")
                pt = r.get("product_type", "")
                excel_row = r.get("excel_row")
                break
        msg = "Удалить результат этой строки? Будет очищена память графа/кэша."
        if price is None and not url:
            msg = "У строки нет результата. Очистить связанную память графа/кэша?"
        ret = QMessageBox.question(self, "Удалить результат", msg,
                                   QMessageBox.Yes | QMessageBox.No)
        if ret != QMessageBox.Yes:
            return
        purged = self._purge_row_memory(spec_text, url=url)
        # Сброс строки таблицы к «пусто»
        t = TOKENS.get(self._current_theme, TOKENS[Theme.DARK])
        for c, default in ((2, ""), (3, "—"), (4, "—"), (5, "—"), (6, ""), (7, "")):
            cell_item = QTableWidgetItem(default)
            cell_item.setForeground(QColor(t["text-muted"]))
            self.results_table.setItem(table_row, c, cell_item)
        # Очистить в _restored_results (сохранить запись-заглушку с ценой None)
        self._restored_results = [
            r for r in self._restored_results if r.get("spec_text") != spec_text
        ]
        # Excel: очистить ячейки результата
        ws = self.excel_writer.ws
        hm = self.excel_writer.header_map
        if ws and hm:
            if excel_row is None:
                for spec in self.excel_writer.get_specs():
                    if spec.text == spec_text:
                        excel_row = spec.row
                        break
            if excel_row:
                for col_key in ("price", "url", "category"):
                    if col_key in hm:
                        ws.cell(excel_row, hm[col_key], None)
        self.add_log("WARN", "user",
                     f"Удалён результат строки «{spec_text[:60]}»"
                     f" (цены: {purged.get('prices', 0)}, подходов: {purged.get('approaches', 0)}, "
                     f"хинтов: {purged.get('hints', 0)})")
        # Удаляем строку из таблицы результатов: позиция вернётся при полном
        # повторном прогоне. Excel очищен выше.
        self.results_table.removeRow(table_row)
        for r in range(self.results_table.rowCount()):
            num_item = QTableWidgetItem(str(r + 1))
            num_item.setTextAlignment(Qt.AlignCenter)
            self.results_table.setItem(r, 0, num_item)
        self._auto_save_session()

    def _purge_row_memory(self, spec_text: str, url: str = "", site: str = "") -> dict:
        """Полная очистка памяти строки: граф (цены/подходы/хинты) + семантик-кэш.

        Вызывается до повторного поиска или при удалении результата.
        """
        from src.memory_manager import MemoryManager
        from src.semantic_cache import SemanticCache
        purged = {"prices": 0, "approaches": 0, "hints": 0}
        try:
            mm = MemoryManager(self._engine)
            purged = mm.purge_row(spec_text, url=url, site_id=site)
        except Exception as e:
            logger.error("Row purge (graph) failed: %s", e)
        try:
            sc = SemanticCache()
            sc.remove_for_row(spec_text, url=url)
        except Exception as e:
            logger.error("Row purge (semantic cache) failed: %s", e)
        return purged

    def _fix_row_type(self, table_row: int, spec_text: str):
        """Переклассификация строки: выбор правильного типа → purge → перепоиск."""
        if self._processing_active:
            return
        products = self._engine.get_all_products()
        if not products:
            QMessageBox.warning(self, "Исправить тип", "Нет доступных типов товаров.")
            return
        # Текущий тип: из результата или по классификации
        current = ""
        for r in self._restored_results:
            if r.get("spec_text") == spec_text:
                current = r.get("product_type", "")
                break
        if not current:
            current = self._engine.classify_product_type(spec_text)

        from gui.reclassify_dialog import ReclassifyDialog
        dlg = ReclassifyDialog(spec_text, current, products, self)
        if not dlg.exec():
            return
        new_type = dlg.selected_type()
        if not new_type or new_type == current:
            return
        if not self._engine.set_product_type_override(spec_text, new_type):
            QMessageBox.warning(self, "Исправить тип", "Не удалось сохранить правило типа.")
            return
        self._purge_row_memory(spec_text)
        self.add_log("WARN", "user",
                     f"Переклассификация «{spec_text[:60]}»: {current or 'unknown'} → {new_type}. "
                     "Запускаю повторный поиск.")
        # После переклассификации строка должна перезапуститься с новым типом.
        self._retry_single_row(table_row, display_type=new_type)

    def _retry_single_row(self, table_row: int, display_type: str = ""):
        """Повторный поиск одной строки (fresh) БЕЗ удаления из таблицы.

        Строка остаётся на месте и помечается «поиск…», чтобы было видно, что
        агент сейчас работает именно над ней. По завершении результат
        обновляется в той же строке (см. _on_retry_row_done).
        """
        if self._processing_active:
            return
        # Get spec_text from the table
        spec_item = self.results_table.item(table_row, 1)
        if not spec_item:
            return
        spec_text = spec_item.text()
        # Find the original SpecItem
        original_spec = None
        for spec in self.excel_writer.get_specs():
            if spec.text == spec_text:
                original_spec = spec
                break
        if not original_spec:
            self.add_log("WARN", "retry", f"Не найден SpecItem для: {spec_text[:60]}")
            return
        # Сброс результата в памяти строки (правило reuse не вернёт старую цену)
        self._restored_results = [
            r for r in self._restored_results
            if r.get("spec_text") != spec_text
        ]
        # Помечаем строку как «идёт повторный поиск» — пользователь видит работу.
        t = TOKENS.get(self._current_theme, TOKENS[Theme.DARK])
        status = "поиск…"
        items = {
            2: display_type or "",
            3: status,
            4: "…",
            5: "…",
            6: "",
            7: "",
        }
        for c, text in items.items():
            cell_item = QTableWidgetItem(text)
            cell_item.setForeground(QColor(t["accent"]))
            cell_item.setToolTip("Идёт повторный поиск этой позиции…")
            self.results_table.setItem(table_row, c, cell_item)
        self.results_table.selectRow(table_row)
        self.results_table.scrollToItem(self.results_table.item(table_row, 1))
        self.config = self._load_config()
        self._processing_active = True
        self._spinner.setFixedSize(20, 20)
        self._spinner.tick()
        self._spinner_timer.start()
        llm_client = llm_providers.create_llm_client(self.config)
        self._retry_row = table_row  # строка, которую сейчас повторяем (для сброса при ошибке)
        # Свежая попытка: снимаем negative-блокировку строки, чтобы runner не
        # вернул «not_found_cached» мгновенно. Остальные кэши не передаём —
        # поиск полностью с нуля.
        retry_caches = None
        try:
            from src.session_cache import NegativeCache
            if self._restored_caches and self._restored_caches.get("negative_cache"):
                nc = NegativeCache()
                nc.from_dict(self._restored_caches["negative_cache"])
                nc.unblock(spec_text)
                retry_caches = {
                    "negative_cache": nc.to_dict(),
                    "site_blacklist": self._restored_caches.get("site_blacklist", {}),
                    "session_facts": self._restored_caches.get("session_facts", {}),
                }
        except Exception:
            retry_caches = None
        self._retry_runner = MCPAgentRunner(
            specs=[original_spec],
            llm_client=llm_client,
            fresh=True,
            use_approaches=self.use_approaches_cb.isChecked(),
            use_site_ranking=self.use_site_ranking_cb.isChecked(),
            ductwork_enabled=self.ductwork_cb.isChecked(),
            skip_registry=self._skip_registry,
            restored_caches=retry_caches,
            restored_results=[],
        )
        self._retry_runner.row_done_signal.connect(
            lambda idx, result: self._on_retry_row_done(table_row, result)
        )
        self._retry_runner.done_signal.connect(self._on_retry_done)
        self._retry_runner.error_signal.connect(self._on_retry_error)
        self._retry_runner.start()
        self.add_log("INFO", "retry", f"Повтор поиска: {spec_text[:60]}")

    def _on_retry_row_done(self, row, result):
        """Обновляет строку результата НА МЕСТЕ после повторного поиска."""
        if row >= self.results_table.rowCount():
            self.results_table.insertRow(self.results_table.rowCount())
        self._restored_results.append(result)
        price = result.get("price")
        conf = result.get("confidence", 0)
        price_text = f"₽{price:,.2f}" if price is not None else "—"
        conf_text = f"{conf:.0%}" if conf else "—"
        elapsed = result.get("elapsed")
        elapsed_text = f"{elapsed:.0f}с" if elapsed is not None else "—"
        site = result.get("site", "")
        url = result.get("url", "")
        spec = result.get("spec_text", "")
        pt = result.get("product_type", "")
        error = result.get("error", "")
        brand_mismatch = result.get("brand_mismatch", False)
        # Номер строки сохраняем прежним (если ячейка есть), иначе row+1
        old_num = self.results_table.item(row, 0)
        num = old_num.text() if old_num is not None else str(row + 1)
        items = [
            num, spec, pt, price_text, conf_text, elapsed_text,
            site if site else "", url if url else ""
        ]
        t = TOKENS.get(self._current_theme, TOKENS[Theme.DARK])
        for c, text in enumerate(items):
            item = QTableWidgetItem(text)
            if c == 7 and url:
                item.setToolTip(url)
                item.setData(Qt.UserRole, url)
            if c == 1 and spec:
                item.setToolTip(spec)
            if error:
                item.setForeground(QColor(t["danger"]))
            elif brand_mismatch:
                item.setForeground(QColor(t["warning"]))
            elif price is not None:
                item.setForeground(QColor(t["success"]))
            else:
                item.setForeground(QColor(t["warning"]))
            self.results_table.setItem(row, c, item)
        # Действия — в контекстном меню строки (ПКМ), кнопок нет.
        self.results_table.scrollToBottom()
        self._retry_row = None
        # Write to Excel
        ws = self.excel_writer.ws
        hm = self.excel_writer.header_map
        if ws and hm:
            excel_row = result.get("excel_row") or (row + 2)
            ws.cell(excel_row, hm["price"], price)
            ws.cell(excel_row, hm["url"], url or "")
            ws.cell(excel_row, hm["category"], pt or "")
        if error:
            self.add_log("WARN", "retry", f"Row {row+1}: {error}")
        elif price:
            self.add_log("INFO", "retry", f"Row {row+1}: {price_text} ({conf_text}) on {site}")

    def _on_retry_done(self, ok, results):
        """Retry runner finished."""
        self._processing_active = False
        self._spinner_timer.stop()
        self._spinner.setFixedSize(0, 0)
        # Если результат строки так и не пришёл (ранний стоп/сбой до row_done),
        # снимаем маркер «поиск…» и возвращаем строку в состояние «не найдено».
        if self._retry_row is not None:
            row = self._retry_row
            self._retry_row = None
            if row < self.results_table.rowCount():
                t = TOKENS.get(self._current_theme, TOKENS[Theme.DARK])
                for c, text in ((3, "—"), (4, "—"), (5, "—"), (6, ""), (7, "")):
                    cell_item = QTableWidgetItem(text)
                    cell_item.setForeground(QColor(t["warning"]))
                    self.results_table.setItem(row, c, cell_item)
        self.add_log("INFO", "retry", "Повтор завершён")

    def _on_retry_error(self, msg):
        """Ошибка retry-runner: сбрасываем маркер «поиск…» строки."""
        self._processing_active = False
        self._spinner_timer.stop()
        self._spinner.setFixedSize(0, 0)
        self.add_log("ERR", "retry", f"Ошибка повтора: {msg}")
        if self._retry_row is not None:
            row = self._retry_row
            self._retry_row = None
            if row < self.results_table.rowCount():
                t = TOKENS.get(self._current_theme, TOKENS[Theme.DARK])
                for c, text in ((3, "—"), (4, "—"), (5, "—"), (6, ""), (7, "")):
                    cell_item = QTableWidgetItem(text)
                    cell_item.setForeground(QColor(t["warning"]))
                    self.results_table.setItem(row, c, cell_item)
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.status_label.setText(f"Ошибка повтора: {msg[:60]}")

    def add_log(self, level, phase, message):
        entry = {
            "timestamp": datetime.now().isoformat(),
            "level": level,
            "phase": phase,
            "message": message,
        }
        self._log_data.append(entry)
        if len(self._log_data) > 5000:
            self._log_data = self._log_data[-2500:]

        self._session_log_entries.append({"level": level, "phase": phase, "msg": str(message)[:200]})
        if len(self._session_log_entries) > 500:
            self._session_log_entries = self._session_log_entries[-300:]

        t = TOKENS.get(self._current_theme, TOKENS[Theme.DARK])
        level_colors = {"INFO": t["success"], "WARN": t["warning"], "ERR": t["danger"], "DEBUG": t["text-muted"]}
        lc = level_colors.get(level, t["text-primary"])
        ts = entry["timestamp"][11:19]
        safe_msg = str(message).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        safe_msg = ui_icons.replace_emojis(safe_msg, px=12)
        border = f'border-left:3px solid {t["danger"]};' if level == "ERR" else \
                 f'border-left:3px solid {t["warning"]};' if level == "WARN" else ''
        self.log_browser.append(
            f'<div style="margin-bottom:2px;line-height:1.3;{border}padding-left:4px;">'
            f'<span style="color:{t["text-muted"]};font-size:11px;">[{ts}]</span>'
            f' <span style="color:{lc};font-weight:600;font-size:11px;">[{level}]</span>'
            f' <span style="color:{t["accent"]};font-size:11px;">[{phase}]</span>'
            f' <span style="color:{t["text-primary"]};font-size:11px;">{safe_msg}</span></div>'
        )
        self.log_browser.verticalScrollBar().setValue(
            self.log_browser.verticalScrollBar().maximum()
        )

    def _toggle_theme(self):
        self._current_theme = Theme.LIGHT if self._current_theme == Theme.DARK else Theme.DARK
        apply_theme(QApplication.instance(), self._current_theme)
        from src.config_loader import save_theme
        save_theme(self._current_theme)
        for widget, key, px in getattr(self, "_icon_specs", []):
            _tt = TOKENS.get(self._current_theme, TOKENS[Theme.DARK])
            try:
                ui_icons.attach(widget, key, _tt["text-primary"], px)
            except Exception:
                pass

    def open_settings(self):
        self.config = self._load_config()
        dlg = SettingsDialog(self.config, self, theme_name=self._current_theme)
        dlg.exec()
        self.config = self._load_config()
        self.assistant_panel.llm_config = self.config.get("llm", {})

    def open_dependency_manager(self):
        from src.dependency_manager.dialog import DependencyManagerDialog
        dlg = DependencyManagerDialog(
            Path(__file__).parent,
            parent=self,
            busy=getattr(self, "_processing_active", False),
        )
        dlg.exec()

    def open_rules_editor(self):
        from gui.rules_editor import RulesEditorDialog
        dlg = RulesEditorDialog(parent=self)
        dlg.exec()

    def load_spec(self, path: str | None = None):
        if not path:
            input_dir = self.config.get("paths", {}).get("data_input", "data/input")
            path, _ = QFileDialog.getOpenFileName(self, "Open spec.xlsx", input_dir, "Excel (*.xlsx)")
            if not path:
                return

        # Архивируем текущую сессию перед загрузкой новой спецификации,
        # чтобы предыдущая сессия не потерялась ( файл _current.json
        # перезаписывается при автосохранении).
        try:
            from src.session_manager import archive_current_session, has_current_session
            if has_current_session() and self._spec_path:
                archive_current_session(self._spec_path)
        except Exception:
            pass

        try:
            headers, data_rows = self.excel_writer.load_spec(path)
            self._spec_path = path
            self._total_rows = data_rows
            self._skip_registry.reset()
            self._restored_results = []
            self._restored_row_indices = set()
            self._restored_caches = None
            self._restored_audit_id = ""
            self._original_restored_results = []
            self._session_log_entries = []
            self.start_btn.setEnabled(True)
            self._show_preview()
            self._center_tabs.setCurrentIndex(1)  # switch to Предпросмотр
            self.add_log("INFO", "init", f"Loaded {self._total_rows} rows from {Path(path).name}")
            mapping = self.excel_writer.detect_columns(headers)
            self.status_label.setText(
                f"Загружено: {self._total_rows} строк · Колонки: {self._mapping_hint(mapping)}"
            )
            self.toast_manager.success(f"Loaded {self._total_rows} rows")
            self._prompt_ductwork_module()
        except Exception as e:
            self.add_log("ERR", "init", f"Failed to load: {e}")
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить файл:\n{e}")

    def _prompt_ductwork_module(self):
        """Детекция воздуховодов в спецификации → дисклаймер «включить модуль?».

        Модуль расчёта воздуховодов по умолчанию ВЫКЛЮЧЕН. При обнаружении
        воздуховодов/фасонных частей предлагается включить его (расчёт без
        обращения к сайтам). Выбор сохраняется в config/settings.yaml.
        """
        from src.config_loader import save_ductwork_enabled
        from src.ductwork_calculator import count_ductwork_items
        try:
            specs = self.excel_writer.get_specs()
        except Exception as e:
            self.add_log("WARN", "init", f"Детекция воздуховодов не выполнена: {e}")
            return
        n = count_ductwork_items(specs)
        if n > 0:
            ans = QMessageBox.question(
                self, "Модуль расчёта воздуховодов",
                f"В спецификации обнаружены воздуховоды и фасонные части — {n} поз.\n\n"
                "Включить модуль расчёта воздуховодов?\n"
                "(позиции будут рассчитаны локально, без обращения к сайтам)\n\n"
                "«Нет» — эти позиции будут обрабатываться как обычно (поиск по сайтам).",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
            )
            enabled = ans == QMessageBox.Yes
            self.ductwork_cb.blockSignals(True)
            self.ductwork_cb.setChecked(enabled)
            self.ductwork_cb.blockSignals(False)
            save_ductwork_enabled(enabled)
            self.add_log("INFO", "init",
                         f"Воздуховоды: обнаружено {n} поз., модуль расчёта {'ВКЛЮЧЁН' if enabled else 'выключен'}")
            self.status_label.setText(
                f"Загружено: {self._total_rows} строк · модуль воздуховодов {'вкл' if enabled else 'выкл'}"
            )
        else:
            save_ductwork_enabled(False)
            self.ductwork_cb.blockSignals(True)
            self.ductwork_cb.setChecked(False)
            self.ductwork_cb.blockSignals(False)

    @staticmethod
    def _mapping_hint(mapping: dict) -> str:
        """Краткий список распознанных колонок для статусной строки."""
        roles = [
            ("name", "Наименование"), ("brand", "Производитель"), ("spec", "Тип/обозначение"),
            ("article", "Артикул"), ("uom", "Ед.изм."), ("qty", "Кол-во"),
        ]
        found = []
        for key, label in roles:
            value = mapping.get(key)
            present = bool(value) if key in ("name", "brand", "spec", "article") else value is not None
            if present:
                found.append(label)
        return ", ".join(found) if found else "нет распознанных колонок"

    def _show_preview(self):
        ws = self.excel_writer.ws
        headers = self.excel_writer.headers
        if ws is None:
            return

        mapping = self.excel_writer.detect_columns(headers)
        name_cols = mapping.get("name", [])
        brand_cols = mapping.get("brand", [])
        spec_cols = mapping.get("spec", [])
        article_cols = mapping.get("article", [])
        qty_col = mapping.get("qty")

        self.preview_table.setRowCount(0)
        self._skip_reconciling = True
        try:
            preview_rows = []
            for excel_row in range(2, ws.max_row + 1):
                name = self.excel_writer.build_item_name(excel_row, mapping)[0]
                if not name or name.strip() in ("", "None", "none"):
                    continue
                spec_item = self.excel_writer.spec_for_row(excel_row)
                preview_rows.append((excel_row, spec_item, name))

            self.preview_table.setRowCount(len(preview_rows))
            for i, (excel_row, spec_item, name) in enumerate(preview_rows):
                check = QTableWidgetItem()
                check.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
                check.setCheckState(Qt.Unchecked)
                check.setData(Qt.UserRole, excel_row)
                check.setData(Qt.UserRole + 1, spec_item.text if spec_item else name)
                check.setData(Qt.UserRole + 2, spec_item.brand if spec_item else "")
                self.preview_table.setItem(i, 0, check)

                self.preview_table.setItem(i, 1, QTableWidgetItem(str(excel_row - 1)))

                self.preview_table.setItem(i, 2, QTableWidgetItem(name[:80]))

                brand = self._concat_display(ws, excel_row, brand_cols)
                self.preview_table.setItem(i, 3, QTableWidgetItem(brand))

                spec = self._concat_display(ws, excel_row, spec_cols)
                self.preview_table.setItem(i, 4, QTableWidgetItem(spec[:80]))

                article_parts = []
                for idx in article_cols:
                    val = str(ws.cell(excel_row, idx + 1).value or "").strip()
                    if val and val not in ("None", ""):
                        article_parts.append(val)
                self.preview_table.setItem(i, 5, QTableWidgetItem(", ".join(article_parts)))

                qty_val = ""
                if qty_col is not None:
                    v = ws.cell(excel_row, qty_col + 1).value
                    if v is not None:
                        qty_val = str(v)
                self.preview_table.setItem(i, 6, QTableWidgetItem(qty_val))
        finally:
            self._skip_reconciling = False
        self._reconcile_skip_checks()

    def _on_preview_item_changed(self, item):
        if self._skip_reconciling or item.column() != 0:
            return
        text = item.data(Qt.UserRole + 1)
        if not text:
            return
        brand = item.data(Qt.UserRole + 2) or ""
        if item.checkState() == Qt.Checked:
            self._skip_registry.mark(text, brand)
        else:
            self._skip_registry.unmark(text, brand)
        self._reconcile_skip_checks()

    def _reconcile_skip_checks(self):
        """Синхронизирует чекбоксы предпросмотра с реестром пропуска (транзитивно)."""
        self._skip_reconciling = True
        try:
            t = TOKENS.get(self._current_theme, TOKENS[Theme.DARK])
            skipped_count = 0
            for row in range(self.preview_table.rowCount()):
                item = self.preview_table.item(row, 0)
                if not item:
                    continue
                text = item.data(Qt.UserRole + 1)
                brand = item.data(Qt.UserRole + 2) or ""
                matched = self._skip_registry.matches(text, brand)
                target = Qt.Checked if matched else Qt.Unchecked
                if item.checkState() != target:
                    item.setCheckState(target)
                item.setToolTip(f"Пропускается: полный аналог «{matched}»" if matched else "")
                if matched:
                    skipped_count += 1
                    fg = QColor(t["text-muted"])
                else:
                    fg = QColor(t["text-primary"])
                for c in range(1, self.preview_table.columnCount()):
                    cell = self.preview_table.item(row, c)
                    if cell:
                        cell.setForeground(fg)
            if hasattr(self, "clear_skip_btn"):
                self.clear_skip_btn.setEnabled(skipped_count > 0)
        finally:
            self._skip_reconciling = False

    def _clear_skip_marks(self):
        self._skip_registry.reset()
        self._reconcile_skip_checks()
        self.add_log("INFO", "init", "Отметки «пропустить» сняты")

    @staticmethod
    def _concat_display(ws, excel_row: int, indices: list[int]) -> str:
        """Собирает значения колонок для предпросмотра (без кавычек бренда)."""
        parts = []
        for idx in indices:
            val = str(ws.cell(excel_row, idx + 1).value or "").strip()
            val = val.strip('"«»')
            if val and val not in ("None", ""):
                parts.append(val)
        return " ".join(parts)

    def _tick_spinner(self):
        self._spinner.tick()

    def start_processing(self):
        if not self._spec_path:
            return

        self.config = self._load_config()
        self._processing_active = True
        self.results_table.setRowCount(0)
        self._restored_results = []
        self.log_browser.clear()
        self.progress_bar.setValue(0)
        self._center_tabs.setCurrentIndex(0)  # switch to Результаты
        self._spinner.setFixedSize(20, 20)
        self._spinner.tick()
        self._spinner_timer.start()

        llm_client = llm_providers.create_llm_client(self.config)

        self._runner = MCPAgentRunner(
            specs=self.excel_writer.get_specs(),
            llm_client=llm_client,
            fresh=not self.reuse_price_cb.isChecked(),
            use_approaches=self.use_approaches_cb.isChecked(),
            use_site_ranking=self.use_site_ranking_cb.isChecked(),
            ductwork_enabled=self.ductwork_cb.isChecked(),
            skip_registry=self._skip_registry,
            restored_caches=self._restored_caches,
            restored_results=self._original_restored_results,
        )
        mode_str = (
            f"цены={'вкл' if self.reuse_price_cb.isChecked() else 'выкл'}, "
            f"подходы={'вкл' if self.use_approaches_cb.isChecked() else 'выкл'}, "
            f"рейтинг={'вкл' if self.use_site_ranking_cb.isChecked() else 'выкл'}, "
            f"воздуховоды={'вкл' if self.ductwork_cb.isChecked() else 'выкл'}"
        )
        self.add_log("INFO", "init", f"Режим поиска: {mode_str}")
        self.monitor_panel.reset()
        self.metrics_panel.reset()
        self._runner.status_signal.connect(self._on_runner_status)
        self._runner.row_done_signal.connect(self._on_row_done)
        self._runner.done_signal.connect(self._on_all_done)
        self._runner.error_signal.connect(self._on_runner_error)
        self._runner.monitor_signal.connect(self._on_monitor_event)
        self._runner.metrics_signal.connect(self._on_metrics)
        self._runner.start()

        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.add_log("INFO", "init", f"Processing {self._total_rows} rows")

    def stop_processing(self):
        if hasattr(self, '_pdf_runner') and self._pdf_runner and self._pdf_runner.isRunning():
            self._pdf_runner.stop()
            self.add_log("INFO", "pdf", "Остановлено пользователем")
            self.status_label.setText("Остановлен")
            return
        if hasattr(self, '_runner') and self._runner:
            self._runner.stop()
        self._spinner_timer.stop()
        self._spinner.setFixedSize(0, 0)
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self._processing_active = False
        self.add_log("INFO", "control", "Остановлено пользователем")
        self.status_label.setText("Остановлен")

    def _on_runner_status(self, status):
        if isinstance(status, tuple):
            _, done, total, msg = status
            self.progress_bar.setMaximum(total)
            self.progress_bar.setValue(done)
            self.status_label.setText(msg)
        elif status == "start":
            self.status_label.setText("Запуск...")
            self.add_log("INFO", "init", "Обработка начата")

    def _on_monitor_event(self, event):
        self.monitor_panel.handle_event(event)

    def _on_metrics(self, stats):
        self.metrics_panel.update_metrics(stats)

    def _on_row_done(self, idx, result):
        self._restored_results.append(result)
        row = self.results_table.rowCount()
        self.results_table.insertRow(row)

        price = result.get("price")
        conf = result.get("confidence", 0)
        price_text = f"₽{price:,.2f}" if price is not None else "—"
        conf_text = f"{conf:.0%}" if conf else "—"
        elapsed = result.get("elapsed")
        elapsed_text = f"{elapsed:.0f}с" if elapsed is not None else "—"
        site = result.get("site", "")
        url = result.get("url", "")
        spec = result.get("spec_text", "")
        pt = result.get("product_type", "")
        error = result.get("error", "")
        brand_mismatch = result.get("brand_mismatch", False)

        items = [
            str(idx + 1), spec, pt, price_text, conf_text, elapsed_text,
            site if site else "", url if url else ""
        ]
        t = TOKENS.get(self._current_theme, TOKENS[Theme.DARK])
        for c, text in enumerate(items):
            item = QTableWidgetItem(text)
            # URL (колонка 7) — ПОЛНЫЙ в ячейке + в UserRole для двойного клика:
            # усечение url[:80] давало битую ссылку (404) — терялся числовой
            # суффикс маркетплейса/карточки (…/103731804).
            if c == 7 and url:
                item.setToolTip(url)
                item.setData(Qt.UserRole, url)
            if c == 1 and spec:
                item.setToolTip(spec)
            if error:
                item.setForeground(QColor(t["danger"]))
            elif brand_mismatch:
                item.setForeground(QColor(t["warning"]))
            elif price is not None:
                item.setForeground(QColor(t["success"]))
            else:
                item.setForeground(QColor(t["warning"]))
            self.results_table.setItem(row, c, item)

        # Действия над строкой — в контекстном меню (ПКМ по строке):
        # повторный поиск, удаление результата, обучение, переклассификация.
        self.results_table.scrollToBottom()

        if error:
            self.add_log("WARN", "agent", f"Row {idx+1}: {error}")
        elif brand_mismatch:
            self.add_log("WARN", "agent", f"Row {idx+1}: {price_text} ({conf_text}) on {site} — НЕ СОВПАДАЕТ БРЕНД")
        elif price:
            self.add_log("INFO", "agent", f"Row {idx+1}: {price_text} ({conf_text}) on {site}")

        ws = self.excel_writer.ws
        hm = self.excel_writer.header_map
        if ws and hm:
            excel_row = result.get("excel_row") or (idx + 2)
            ws.cell(excel_row, hm["price"], price)
            ws.cell(excel_row, hm["url"], url or "")
            ws.cell(excel_row, hm["category"], pt or "")
            if brand_mismatch and "note" in hm:
                ws.cell(excel_row, hm["note"], "не совпадает бренд")
            elif result.get("ductwork_breakdown") and "note" in hm:
                ws.cell(excel_row, hm["note"], result["ductwork_breakdown"])

    def _switch_to_assistant(self):
        """Switch right panel to assistant tab and ensure it's visible."""
        for i in range(self._right_tabs.count()):
            if self._right_tabs.tabText(i) == "Ассистент":
                self._right_tabs.setCurrentIndex(i)
                break
        sizes = self._splitter.sizes()
        if sizes[1] < 100:
            self._splitter.setSizes([700, 500])

    def _open_study(self, spec_text: str, product_type: str, failure_context: str = ""):
        """Open study tool in assistant panel pre-filled with row data."""
        self.assistant_panel.prefill_study(spec_text, product_type, failure_context)
        self._switch_to_assistant()

    def _open_study_tool(self):
        """Open study tool with empty fields (user fills manually)."""
        self.assistant_panel.prefill_study("", "")
        self._switch_to_assistant()

    def _load_pdf(self):
        if getattr(self, "_pdf_runner", None) and self._pdf_runner.isRunning():
            self.add_log("WARN", "pdf", "PDF уже обрабатывается")
            return

        path, _ = QFileDialog.getOpenFileName(self, "Открыть PDF спецификацию", "",
                                               "PDF (*.pdf)")
        if not path:
            return

        self.pdf_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self._spinner_timer.start()
        self._spinner.setFixedSize(20, 20)

        llm_client = llm_providers.create_llm_client(self.config, temperature=0.1)

        pipeline = self.config.get('pdf_parser', {}).get('pipeline', 'legacy')
        if pipeline == 'v2' and _HAS_V2:
            self._pdf_runner = Pdf2SpecRunner(
                pdf_path=path,
                llm_client=llm_client,
                config=self.config,
            )
        else:
            self._pdf_runner = PdfParserRunner(
                pdf_path=path,
                llm_client=llm_client,
                config=self.config,
            )
        self._pdf_runner.progress_signal.connect(self._on_pdf_progress)
        self._pdf_runner.items_ready_signal.connect(self._on_pdf_items_ready)
        self._pdf_runner.done_signal.connect(self._on_pdf_done)
        self._pdf_runner.start()

        self.add_log("INFO", "pdf", f"Загрузка PDF: {Path(path).name}")

    def _on_pdf_progress(self, msg: str, step: int, total: int):
        self.status_label.setText(msg)
        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(step)

    def _on_pdf_items_ready(self, items: list[dict]):
        self.add_log("INFO", "pdf", f"Извлечено {len(items)} позиций из PDF")

        from src.pdf_parser.review import SmartReview
        auto = sum(1 for it in items if (it.get("confidence") or 0) >= SmartReview.CONFIDENCE_THRESHOLD)
        if auto:
            self.add_log("INFO", "pdf", f"SmartReview: {auto} авто-подтверждено, {len(items) - auto} требует проверки")

        dlg = ReviewDialog(items, self, theme_name=self._current_theme)
        if dlg.exec() and dlg.is_confirmed:
            confirmed = dlg.items
            self.add_log("INFO", "pdf", f"Подтверждено {len(confirmed)} позиций")

            feedback = FeedbackCollector()
            for orig, cor in zip(items, confirmed):
                orig_name = orig.get("name", "")
                cor_name = cor.get("name", "")
                if orig_name != cor_name:
                    feedback.save_correction(orig_name, cor_name, "manual")
                    self.add_log("INFO", "pdf", f"Исправление: «{orig_name}» → «{cor_name}»")

            xlsx_path = self._save_pdf_items_to_excel(confirmed)
            if xlsx_path:
                self.load_spec(path=xlsx_path)
                self.toast_manager.success(f"PDF: {len(confirmed)} позиций готово к старту")
        else:
            self.add_log("INFO", "pdf", "PDF спецификация отклонена пользователем")

    def _build_spec_text(self, item: dict) -> str:
        pos = item.get("pos", 0)
        name = item.get("name", "")
        specs = item.get("specs", "")
        qty = item.get("qty", 0)
        unit = item.get("unit", "шт")
        parts = [f"{pos}. {name}"]
        if specs:
            parts.append(f" {specs}")
        parts.append(f" — {qty} {unit}")
        return "".join(parts)

    def _load_pdf_item_into_spec(self, spec_text: str):
        row = self.results_table.rowCount()
        self.results_table.insertRow(row)
        t = TOKENS.get(self._current_theme, TOKENS[Theme.DARK])
        items = [
            str(row + 1),
            spec_text[:80] if spec_text else "",
            "PDF",
            "",
            "",
            "",
            "",
            "",
        ]
        for c, text in enumerate(items):
            item = QTableWidgetItem(text)
            item.setForeground(QColor(t["accent"]))
            self.results_table.setItem(row, c, item)

    def _save_pdf_items_to_excel(self, items: list[dict]):
        out_dir = self.config.get("paths", {}).get("data_output", "data/output")
        Path(out_dir).mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = Path(out_dir) / f"pdf_spec_{ts}.xlsx"

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PDF спецификация"
        ws.append(["#", "Наименование", "Характеристики", "Артикул", "Изготовитель", "Кол-во", "Ед."])
        for item in items:
            ws.append([
                item.get("pos", ""),
                item.get("name", ""),
                item.get("specs", ""),
                item.get("code", ""),
                item.get("manufacturer", ""),
                item.get("qty", ""),
                item.get("unit", "шт"),
            ])
        wb.save(str(out_path))
        self.add_log("INFO", "pdf", f"Сохранено: {out_path.name}")
        return str(out_path)

    def _on_pdf_done(self, success: bool, msg: str):
        self._spinner_timer.stop()
        self._spinner.setFixedSize(0, 0)
        self.pdf_btn.setEnabled(True)
        if not self._processing_active:
            self.stop_btn.setEnabled(False)
        self.status_label.setText(msg if success else f"Ошибка: {msg}")
        if success:
            self.add_log("INFO", "pdf", msg)
        else:
            self.add_log("ERR", "pdf", msg)

    @staticmethod
    def _norm(text: str) -> str:
        """Нормализация для сравнения: lowercase + схлопывание пробелов."""
        return " ".join((text or "").lower().split())

    def _merge_session_results(self, runner_results: list) -> list:
        """Merge runner results with original restored results.

        Core rule: iterate over _original_restored_results (ALL specs from session)
        and supplement with runner_results (latest data). This ensures no spec is
        ever lost, even if the runner was stopped or crashed mid-way.
        """
        _old_by_spec = {r.get("spec_text"): r for r in self._original_restored_results
                        if r.get("spec_text")}
        _runner_by_spec = {r.get("spec_text"): r for r in runner_results
                           if r.get("spec_text")}
        # Normalized fallback index (lowercase + collapse whitespace)
        _old_by_norm = {}
        for r in self._original_restored_results:
            st = r.get("spec_text", "")
            if st:
                _old_by_norm[self._norm(st)] = r
        _runner_by_norm = {}
        for r in runner_results:
            st = r.get("spec_text", "")
            if st:
                _runner_by_norm[self._norm(st)] = r
        _new_results = []
        # 1. Walk ALL original results — keep or upgrade each
        for r in self._original_restored_results:
            st = r.get("spec_text", "")
            # Exact match first, then normalized fallback
            runner_r = _runner_by_spec.get(st)
            if runner_r is None and st:
                runner_r = _runner_by_norm.get(self._norm(st))
            if runner_r:
                if runner_r.get("price") is not None:
                    _new_results.append(runner_r)       # new result with price → use it
                elif r.get("price") is not None:
                    _new_results.append(r)               # old has price, new doesn't → keep old
                else:
                    _new_results.append(runner_r)         # both no price → use new (fresh attempt)
            else:
                _new_results.append(r)                    # runner didn't process → keep old
        # 2. Add any NEW specs from runner not in original list
        _old_specs = {r.get("spec_text") for r in self._original_restored_results}
        _old_norms = {self._norm(s) for s in _old_specs if s}
        for r in runner_results:
            st = r.get("spec_text", "")
            if st and st not in _old_specs and self._norm(st) not in _old_norms:
                _new_results.append(r)
        return _new_results

    def _on_all_done(self, success, spec_result):
        self._spinner_timer.stop()
        self._spinner.setFixedSize(0, 0)
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self._processing_active = False

        total = spec_result.get("total", 0)
        found = spec_result.get("found_count", 0)
        errs = spec_result.get("error_count", 0)

        if hasattr(self, '_runner') and self._runner and self._runner.results:
            _new_results = self._merge_session_results(self._runner.results)
            self._restored_results = _new_results

        self.add_log("INFO", "complete",
            f"Готово: {found}/{total} найдено, {errs} ошибок")
        self.status_label.setText(f"Готово: {found}/{total}")

        if found > 0:
            self.toast_manager.success(f"Found {found}/{total} prices")
        else:
            self.toast_manager.warning(f"No prices found ({errs} errors)")

        if self._spec_path:
            try:
                out_dir = self.config.get("paths", {}).get("data_output", "data/output")
                out_path = self.excel_writer.save_output_copy(out_dir)
                self.add_log("INFO", "output", f"Saved to {Path(out_path).name}")
            except Exception as e:
                self.add_log("ERR", "output", f"Save failed: {e}")

    def _on_runner_error(self, msg):
        self._spinner_timer.stop()
        self._spinner.setFixedSize(0, 0)
        self.add_log("ERR", "runner", msg)
        # Merge partial runner results so nothing is lost on crash
        if hasattr(self, '_runner') and self._runner and self._runner.results:
            _new_results = self._merge_session_results(self._runner.results)
            self._restored_results = _new_results
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self._processing_active = False
        self.status_label.setText(f"Ошибка: {msg[:60]}")

    def _on_headless_toggle(self, checked):
        from src.config_loader import save_browser_headless
        save_browser_headless(checked)
        if self._processing_active and hasattr(self, '_runner') and self._runner:
            self._runner.trigger_bridge_restart(checked)
            self.add_log("INFO", "control", f"Bridge restarting with headless={checked}")

    def _on_backend_change(self, index):
        backend = self.backend_combo.itemData(index)
        if not backend:
            return
        from src.config_loader import save_browser_backend
        save_browser_backend(backend)
        if self._processing_active and hasattr(self, '_runner') and self._runner:
            self._runner.trigger_bridge_backend_restart(backend)
            self.add_log("INFO", "control", f"Bridge restarting with backend={backend}")

    def _on_run_mode_toggle(self, checked=False):
        from src.config_loader import save_run_flags, save_ductwork_enabled
        reuse_price = self.reuse_price_cb.isChecked()
        use_approaches = self.use_approaches_cb.isChecked()
        use_site_ranking = self.use_site_ranking_cb.isChecked()
        save_run_flags(reuse_price=reuse_price, use_approaches=use_approaches,
                       use_site_ranking=use_site_ranking)
        save_ductwork_enabled(self.ductwork_cb.isChecked())
        if self._processing_active and hasattr(self, '_runner') and self._runner:
            self._runner.set_fresh(not reuse_price)
            self._runner.set_use_approaches(use_approaches)
            self._runner.set_use_site_ranking(use_site_ranking)
            self._runner.set_ductwork_enabled(self.ductwork_cb.isChecked())
            mode = (f"цены={'вкл' if reuse_price else 'выкл'}, "
                    f"подходы={'вкл' if use_approaches else 'выкл'}, "
                    f"рейтинг={'вкл' if use_site_ranking else 'выкл'}, "
                    f"воздуховоды={'вкл' if self.ductwork_cb.isChecked() else 'выкл'}")
            self.add_log("INFO", "control", f"Режим поиска: {mode} (со следующей позиции)")


def main():
    from PySide6.QtGui import QSurfaceFormat
    fmt = QSurfaceFormat()
    fmt.setSamples(8)
    fmt.setSwapInterval(1)
    QSurfaceFormat.setDefaultFormat(fmt)

    app = QApplication(sys.argv)
    app.setStyle(QStyleFactory.create("Fusion"))
    window = MainWindow()
    window.showMaximized()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
